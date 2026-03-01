"""Operation context and reference resolution for sheets verbs.

Provides SheetsOpContext and helpers for resolving cell references,
anchors (C6), and selectors.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from typing import Iterator

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from fcp_sheets.model.index import SheetIndex
from fcp_sheets.model.refs import (
    AnchorRef,
    CellRef,
    ColRef,
    RangeRef,
    RowRef,
    col_to_index,
    index_to_col,
    parse_anchor,
    parse_cell_ref,
    parse_range_ref,
)


@dataclass
class SheetsOpContext:
    """Context passed to every verb handler."""

    wb: Workbook
    index: SheetIndex
    named_styles: dict[str, dict]

    @property
    def active_sheet(self) -> Worksheet:
        """The currently active worksheet.

        Uses name-based tracking from index as primary source of truth,
        falling back to openpyxl's index-based wb.active if the name
        is missing or stale.
        """
        name = self.index.active_sheet
        if name and name in self.wb.sheetnames:
            return self.wb[name]
        return self.wb.active  # type: ignore[return-value]

    @property
    def active_sheet_name(self) -> str:
        """Name of the active worksheet."""
        name = self.index.active_sheet
        if name and name in self.wb.sheetnames:
            return name
        ws = self.wb.active
        return ws.title if ws else ""


def resolve_cell_ref(ref_str: str, ctx: SheetsOpContext) -> tuple[int, int] | None:
    """Resolve a cell reference string to (col, row).

    Handles standard A1 refs AND C6 spatial anchors.
    Returns None if ref cannot be parsed.
    """
    ref_str = ref_str.strip()

    # Try anchor first (C6)
    anchor = parse_anchor(ref_str)
    if anchor:
        return resolve_anchor(anchor, ctx)

    # Standard A1 ref
    cell = parse_cell_ref(ref_str)
    if cell:
        return (cell.col, cell.row)

    return None


def resolve_anchor(anchor: AnchorRef, ctx: SheetsOpContext) -> tuple[int, int] | None:
    """Resolve a spatial anchor to (col, row) based on current data bounds (C6).

    Anchors:
      @bottom_left   → (min_col, max_row + 1 + offset)
      @bottom_right  → (max_col, max_row + 1 + offset)
      @right_top     → (max_col + 1 + offset, min_row)
    """
    sheet_name = ctx.active_sheet_name
    bounds = ctx.index.get_bounds(sheet_name)

    if bounds is None:
        # No data — treat as (1, 1) + offset
        if anchor.anchor in ("bottom_left", "bottom_right"):
            return (1, 1 + anchor.offset)
        else:  # right_top
            return (1 + anchor.offset, 1)

    min_row, min_col, max_row, max_col = bounds

    if anchor.anchor == "bottom_left":
        return (min_col, max_row + 1 + anchor.offset)
    elif anchor.anchor == "bottom_right":
        return (max_col, max_row + 1 + anchor.offset)
    elif anchor.anchor == "right_top":
        return (max_col + 1 + anchor.offset, min_row)

    return None


def get_target_sheet(ref_str: str, ctx: SheetsOpContext) -> tuple[Worksheet, str]:
    """Extract target sheet from a cross-sheet reference.

    Returns (worksheet, ref_without_sheet_prefix).
    """
    if "!" in ref_str:
        sheet_name, _, rest = ref_str.partition("!")
        sheet_name = sheet_name.strip("'\"")
        if sheet_name in ctx.wb.sheetnames:
            return ctx.wb[sheet_name], rest
    return ctx.active_sheet, ref_str


def resolve_range_to_cells(
    range_str: str, ctx: SheetsOpContext
) -> Iterator[tuple[Worksheet, int, int]]:
    """Resolve a range string to an iterator of (worksheet, row, col) tuples."""
    ws, ref = get_target_sheet(range_str, ctx)

    # Single cell
    cell = parse_cell_ref(ref)
    if cell:
        yield (ws, cell.row, cell.col)
        return

    # Range
    range_ref = parse_range_ref(ref)
    if isinstance(range_ref, RangeRef):
        for row in range(range_ref.start.row, range_ref.end.row + 1):
            for col in range(range_ref.start.col, range_ref.end.col + 1):
                yield (ws, row, col)
        return

    if isinstance(range_ref, ColRef):
        # Use data bounds to limit iteration
        sheet_name = ws.title
        bounds = ctx.index.get_bounds(sheet_name)
        if bounds:
            min_row, _, max_row, _ = bounds
            for row in range(min_row, max_row + 1):
                for col in range(range_ref.start_col, range_ref.end_col + 1):
                    yield (ws, row, col)
        return

    if isinstance(range_ref, RowRef):
        bounds = ctx.index.get_bounds(ws.title)
        if bounds:
            _, min_col, _, max_col = bounds
            for row in range(range_ref.start_row, range_ref.end_row + 1):
                for col in range(min_col, max_col + 1):
                    yield (ws, row, col)
        return


# ---------------------------------------------------------------------------
# Selector resolution
# ---------------------------------------------------------------------------

def resolve_selectors(
    selectors: list[str], ctx: SheetsOpContext
) -> list[tuple[Worksheet, int, int]]:
    """Resolve a list of selector strings to a list of (worksheet, row, col).

    Multiple selectors are intersected (AND logic).
    Scans within data_bounds per C4 (lazy evaluation).

    Supported selectors:
      @sheet:NAME       — cells on named sheet
      @range:A1:Z99     — cells in range
      @row:N  / @row:N-M
      @col:A  / @col:A-E
      @type:formula|number|text|date|empty
      @table:NAME       — cells in named table
      @name:NAME        — cells in named range
      @all              — all cells with data (within data_bounds)
      @recent / @recent:N
      @not:TYPE:VALUE   — negated selector
    """
    if not selectors:
        return []

    result_sets: list[set[tuple[str, int, int]]] = []

    for sel in selectors:
        cells = _resolve_single_selector(sel, ctx)
        result_sets.append(cells)

    # Intersect all sets
    if not result_sets:
        return []

    combined = result_sets[0]
    for s in result_sets[1:]:
        combined = combined & s

    # Convert (sheet_name, row, col) back to (Worksheet, row, col)
    result: list[tuple[Worksheet, int, int]] = []
    for sheet_name, row, col in sorted(combined):
        if sheet_name in ctx.wb.sheetnames:
            ws = ctx.wb[sheet_name]
            result.append((ws, row, col))

    return result


def _resolve_single_selector(
    sel: str, ctx: SheetsOpContext
) -> set[tuple[str, int, int]]:
    """Resolve a single selector string to a set of (sheet_name, row, col)."""
    sel = sel.strip()

    # @not:TYPE:VALUE — negated selector
    if sel.startswith("@not:"):
        inner = "@" + sel[5:]  # e.g., "@not:type:formula" -> "@type:formula"
        # Get the universe (all cells with data)
        universe = _resolve_all(ctx)
        excluded = _resolve_single_selector(inner, ctx)
        return universe - excluded

    # @all
    if sel == "@all":
        return _resolve_all(ctx)

    # @recent / @recent:N
    if sel == "@recent" or sel.startswith("@recent:"):
        return _resolve_recent(sel, ctx)

    # @sheet:NAME
    if sel.startswith("@sheet:"):
        sheet_name = sel[7:]
        return _resolve_sheet(sheet_name, ctx)

    # @range:A1:Z99
    if sel.startswith("@range:"):
        range_str = sel[7:]
        return _resolve_range(range_str, ctx)

    # @row:N or @row:N-M
    if sel.startswith("@row:"):
        row_spec = sel[5:]
        return _resolve_row(row_spec, ctx)

    # @col:A or @col:A-E
    if sel.startswith("@col:"):
        col_spec = sel[5:]
        return _resolve_col(col_spec, ctx)

    # @type:formula|number|text|date|empty
    if sel.startswith("@type:"):
        type_name = sel[6:]
        return _resolve_type(type_name, ctx)

    # @table:NAME
    if sel.startswith("@table:"):
        table_name = sel[7:]
        return _resolve_table(table_name, ctx)

    # @name:NAME
    if sel.startswith("@name:"):
        named_range = sel[6:]
        return _resolve_named_range(named_range, ctx)

    return set()


def _resolve_all(ctx: SheetsOpContext) -> set[tuple[str, int, int]]:
    """Resolve @all — all cells within data_bounds on active sheet."""
    ws = ctx.active_sheet
    sheet_name = ws.title
    bounds = ctx.index.get_bounds(sheet_name)
    if not bounds:
        return set()
    min_row, min_col, max_row, max_col = bounds
    cells: set[tuple[str, int, int]] = set()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                cells.add((sheet_name, row, col))
    return cells


def _resolve_recent(sel: str, ctx: SheetsOpContext) -> set[tuple[str, int, int]]:
    """Resolve @recent or @recent:N."""
    count = 1
    if ":" in sel and sel != "@recent":
        try:
            count = int(sel.split(":")[1])
        except (ValueError, IndexError):
            count = 1
    count = max(1, count)

    recent = ctx.index.get_recent(count)
    cells: set[tuple[str, int, int]] = set()
    for sheet_name, range_str in recent:
        if sheet_name not in ctx.wb.sheetnames:
            continue
        # Parse the range_str — could be "A1" or "A1..D5"
        # The index stores ranges as "A1..D5" or single "A1"
        range_str = range_str.replace("..", ":")
        ref = parse_cell_ref(range_str)
        if ref:
            cells.add((sheet_name, ref.row, ref.col))
            continue
        rr = parse_range_ref(range_str)
        if isinstance(rr, RangeRef):
            for row in range(rr.start.row, rr.end.row + 1):
                for col in range(rr.start.col, rr.end.col + 1):
                    cells.add((sheet_name, row, col))
    return cells


def _resolve_sheet(sheet_name: str, ctx: SheetsOpContext) -> set[tuple[str, int, int]]:
    """Resolve @sheet:NAME — all cells with data on named sheet."""
    if sheet_name not in ctx.wb.sheetnames:
        return set()
    ws = ctx.wb[sheet_name]
    bounds = ctx.index.get_bounds(sheet_name)
    if not bounds:
        return set()
    min_row, min_col, max_row, max_col = bounds
    cells: set[tuple[str, int, int]] = set()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                cells.add((sheet_name, row, col))
    return cells


def _resolve_range(range_str: str, ctx: SheetsOpContext) -> set[tuple[str, int, int]]:
    """Resolve @range:A1:D10 — cells in a range (on active sheet unless prefixed)."""
    cells: set[tuple[str, int, int]] = set()
    for ws, row, col in resolve_range_to_cells(range_str, ctx):
        cells.add((ws.title, row, col))
    return cells


def _resolve_row(row_spec: str, ctx: SheetsOpContext) -> set[tuple[str, int, int]]:
    """Resolve @row:N or @row:N-M."""
    ws = ctx.active_sheet
    sheet_name = ws.title
    bounds = ctx.index.get_bounds(sheet_name)
    if not bounds:
        return set()
    _, min_col, _, max_col = bounds

    if "-" in row_spec:
        parts = row_spec.split("-", 1)
        start_row = int(parts[0])
        end_row = int(parts[1])
    else:
        start_row = end_row = int(row_spec)

    cells: set[tuple[str, int, int]] = set()
    for row in range(start_row, end_row + 1):
        for col in range(min_col, max_col + 1):
            cells.add((sheet_name, row, col))
    return cells


def _resolve_col(col_spec: str, ctx: SheetsOpContext) -> set[tuple[str, int, int]]:
    """Resolve @col:A or @col:A-E."""
    ws = ctx.active_sheet
    sheet_name = ws.title
    bounds = ctx.index.get_bounds(sheet_name)
    if not bounds:
        return set()
    min_row, _, max_row, _ = bounds

    if "-" in col_spec:
        parts = col_spec.split("-", 1)
        start_col = col_to_index(parts[0])
        end_col = col_to_index(parts[1])
    else:
        start_col = end_col = col_to_index(col_spec)

    cells: set[tuple[str, int, int]] = set()
    for row in range(min_row, max_row + 1):
        for col in range(start_col, end_col + 1):
            cells.add((sheet_name, row, col))
    return cells


def _resolve_type(
    type_name: str, ctx: SheetsOpContext
) -> set[tuple[str, int, int]]:
    """Resolve @type:formula|number|text|date|empty — lazy scan within data_bounds."""
    ws = ctx.active_sheet
    sheet_name = ws.title
    bounds = ctx.index.get_bounds(sheet_name)
    if not bounds:
        return set()
    min_row, min_col, max_row, max_col = bounds
    cells: set[tuple[str, int, int]] = set()
    type_name = type_name.lower()

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            match = False

            if type_name == "formula":
                match = isinstance(val, str) and val.startswith("=")
            elif type_name == "number":
                match = isinstance(val, (int, float)) and not isinstance(val, bool)
            elif type_name == "text":
                match = isinstance(val, str) and not val.startswith("=")
            elif type_name == "date":
                match = isinstance(val, datetime)
            elif type_name == "empty":
                match = val is None

            if match:
                cells.add((sheet_name, row, col))
    return cells


def _resolve_table(
    table_name: str, ctx: SheetsOpContext
) -> set[tuple[str, int, int]]:
    """Resolve @table:NAME — cells in a named table."""
    cells: set[tuple[str, int, int]] = set()
    for ws in ctx.wb.worksheets:
        for table in ws.tables.values():
            if table.name == table_name or table.displayName == table_name:
                ref = parse_range_ref(table.ref)
                if isinstance(ref, RangeRef):
                    for row in range(ref.start.row, ref.end.row + 1):
                        for col in range(ref.start.col, ref.end.col + 1):
                            cells.add((ws.title, row, col))
                return cells
    return cells


def _resolve_named_range(
    name: str, ctx: SheetsOpContext
) -> set[tuple[str, int, int]]:
    """Resolve @name:NAME — cells in a named range (defined name)."""
    cells: set[tuple[str, int, int]] = set()
    if name in ctx.wb.defined_names:
        defn = ctx.wb.defined_names[name]
        for title, coord in defn.destinations:
            if title in ctx.wb.sheetnames:
                # coord is like "A1:D10" or "A1"
                ref = parse_cell_ref(coord)
                if ref:
                    cells.add((title, ref.row, ref.col))
                    continue
                rr = parse_range_ref(coord)
                if isinstance(rr, RangeRef):
                    for row in range(rr.start.row, rr.end.row + 1):
                        for col in range(rr.start.col, rr.end.col + 1):
                            cells.add((title, row, col))
    return cells


def resolve_target_cells(
    op_positionals: list[str], op_selectors: list[str], ctx: SheetsOpContext
) -> list[tuple[Worksheet, int, int]]:
    """Resolve target cells from either the first positional (range) or selectors.

    If selectors are present, use resolve_selectors.
    Otherwise, parse the first positional as a range reference.
    Returns a list of (Worksheet, row, col).
    """
    # Selectors take priority
    if op_selectors:
        return resolve_selectors(op_selectors, ctx)

    # Check if first positional looks like a selector
    if op_positionals and op_positionals[0].startswith("@"):
        # Could be a selector passed as positional
        sel = op_positionals[0]
        # Skip spatial anchors (e.g. @bottom_left)
        if not parse_anchor(sel):
            return resolve_selectors([sel], ctx)

    # Fall back to range parsing from first positional
    if not op_positionals:
        return []

    return list(resolve_range_to_cells(op_positionals[0], ctx))
