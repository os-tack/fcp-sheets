"""Cell operation handlers — set, data, fill, clear."""

from __future__ import annotations

from openpyxl.cell.cell import MergedCell

from fcp_core import OpResult, ParsedOp

from fcp_sheets.lib.number_formats import resolve_format
from fcp_sheets.model.refs import col_to_index, index_to_col, parse_cell_ref
from fcp_sheets.server.resolvers import (
    SheetsOpContext,
    resolve_cell_ref,
    resolve_range_to_cells,
)


def op_set(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Set a single cell value.

    Syntax: set CELL VALUE [fmt:FORMAT]

    Values starting with = are formulas.
    Numeric strings are converted to numbers.
    Quoted strings stay as text.
    """
    if len(op.positionals) < 2:
        return OpResult(success=False, message="Usage: set CELL VALUE [fmt:FORMAT]")

    ref_str = op.positionals[0]
    value_str = op.positionals[1]

    # Resolve cell reference (supports A1 and C6 anchors)
    resolved = resolve_cell_ref(ref_str, ctx)
    if resolved is None:
        return OpResult(success=False, message=f"Invalid cell reference: {ref_str!r}")

    col, row = resolved
    ws = ctx.active_sheet

    # Check for merged cell — writing to non-top-left cells in a merged range
    # raises AttributeError in openpyxl
    existing = ws.cell(row=row, column=col)
    if isinstance(existing, MergedCell):
        addr = f"{index_to_col(col)}{row}"
        return OpResult(
            success=False,
            message=f"Cannot write to {addr}: cell is part of a merged range. Write to the top-left cell instead.",
        )

    # Parse value
    value = _parse_cell_value(value_str)

    # Set the cell
    cell = ws.cell(row=row, column=col, value=value)

    # Apply number format if specified
    fmt = op.params.get("fmt")
    if fmt:
        cell.number_format = resolve_format(fmt)

    # Update index bounds
    ctx.index.expand_bounds(ws.title, row, col)
    addr = f"{index_to_col(col)}{row}"
    ctx.index.record_modified(ws.title, addr)

    # Format response
    display = repr(value) if isinstance(value, str) and not value.startswith("=") else str(value)
    return OpResult(success=True, message=f"{addr} = {display}", prefix="+")


def op_data(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Data block mode — handled by adapter block buffering, not dispatched here."""
    raise NotImplementedError("data verb not yet implemented")


def op_fill(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Fill/drag cell value or formula.

    Syntax:
      fill SRC dir:down|right [count:N | to:CELL] [until:COL]

    Reads the source cell and fills adjacent cells in the given direction.
    For formulas, uses openpyxl's Translator for safe reference shifting (C3):
    - $A$1 absolute refs stay locked
    - A1 relative refs shift with the fill direction
    - Cross-sheet refs (Sheet2!$B$2) are handled
    - On Translator failure, copies formula as-is + warns
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: fill SRC dir:down|right [count:N|to:CELL] [until:COL]")

    src_str = op.positionals[0]
    direction = op.params.get("dir", "").lower()
    if direction not in ("down", "right"):
        return OpResult(success=False, message="fill requires dir:down or dir:right")

    # Resolve source cell
    resolved = resolve_cell_ref(src_str, ctx)
    if resolved is None:
        return OpResult(success=False, message=f"Invalid source cell: {src_str!r}")

    src_col, src_row = resolved
    ws = ctx.active_sheet

    # Read source cell value
    src_cell = ws.cell(row=src_row, column=src_col)
    src_value = src_cell.value
    if src_value is None:
        return OpResult(success=False, message=f"Source cell {index_to_col(src_col)}{src_row} is empty")

    is_formula = isinstance(src_value, str) and src_value.startswith("=")

    # Determine target cells based on params
    count = op.params.get("count")
    to_cell = op.params.get("to")
    until_col = op.params.get("until")

    if count is not None:
        try:
            n = int(count)
        except ValueError:
            return OpResult(success=False, message=f"Invalid count: {count!r}")
        if n < 1:
            return OpResult(success=False, message="count must be >= 1")

        if direction == "down":
            targets = [(src_col, src_row + i) for i in range(1, n + 1)]
        else:  # right
            targets = [(src_col + i, src_row) for i in range(1, n + 1)]

    elif to_cell is not None:
        to_resolved = resolve_cell_ref(to_cell, ctx)
        if to_resolved is None:
            return OpResult(success=False, message=f"Invalid target cell: {to_cell!r}")
        to_col, to_row = to_resolved

        if direction == "down":
            if to_row <= src_row:
                return OpResult(success=False, message="to: cell must be below source for dir:down")
            targets = [(src_col, r) for r in range(src_row + 1, to_row + 1)]
        else:  # right
            if to_col <= src_col:
                return OpResult(success=False, message="to: cell must be right of source for dir:right")
            targets = [(c, src_row) for c in range(src_col + 1, to_col + 1)]

    elif until_col is not None:
        # Fill down until the specified column is empty in the target row
        check_col_idx = col_to_index(until_col.upper()) if until_col.isalpha() else int(until_col)
        targets = []
        r = src_row + 1
        max_scan = 10000  # safety limit
        while r - src_row <= max_scan:
            check_val = ws.cell(row=r, column=check_col_idx).value
            if check_val is None or check_val == "":
                break
            targets.append((src_col, r))
            r += 1
    else:
        return OpResult(success=False, message="fill requires count:N, to:CELL, or until:COL")

    if not targets:
        return OpResult(success=True, message="No cells to fill", prefix="~")

    # Fill each target
    src_addr = f"{index_to_col(src_col)}{src_row}"
    warnings: list[str] = []
    filled = 0

    for tgt_col, tgt_row in targets:
        tgt_addr = f"{index_to_col(tgt_col)}{tgt_row}"

        # Skip merged cells (non-top-left cells in a merged range)
        if isinstance(ws.cell(row=tgt_row, column=tgt_col), MergedCell):
            warnings.append(f"! Skipped {tgt_addr}: part of merged range")
            continue

        if is_formula:
            # C3: Safe formula translation using openpyxl Translator
            try:
                from openpyxl.formula.translate import Translator
                translated = Translator(src_value, origin=src_addr).translate_formula(tgt_addr)
                ws.cell(row=tgt_row, column=tgt_col, value=translated)
            except Exception:
                # Fallback: copy formula as-is
                ws.cell(row=tgt_row, column=tgt_col, value=src_value)
                if not warnings:
                    warnings.append("! Warning: Formula translation failed, copied as-is")
        else:
            ws.cell(row=tgt_row, column=tgt_col, value=src_value)

        ctx.index.expand_bounds(ws.title, tgt_row, tgt_col)
        filled += 1

    # Record modified range
    first_tgt = targets[0]
    last_tgt = targets[-1]
    range_str = f"{index_to_col(first_tgt[0])}{first_tgt[1]}..{index_to_col(last_tgt[0])}{last_tgt[1]}"
    ctx.index.record_modified(ws.title, range_str)

    msg = f"Filled {filled} cells from {src_addr} {direction}"
    if warnings:
        msg += "\n" + "\n".join(warnings)
    return OpResult(success=True, message=msg, prefix="+")


def op_clear(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Clear cell contents and optionally formatting.

    Syntax:
      clear RANGE           — clear values only
      clear RANGE all       — clear values AND formatting
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: clear RANGE [all]")

    range_str = op.positionals[0]
    clear_all = len(op.positionals) > 1 and op.positionals[1].lower() == "all"

    # Resolve range to cells
    cells = list(resolve_range_to_cells(range_str, ctx))
    if not cells:
        return OpResult(success=False, message=f"Invalid or empty range: {range_str!r}")

    cleared = 0
    for ws, row, col in cells:
        cell = ws.cell(row=row, column=col)
        cell.value = None

        if clear_all:
            # Reset formatting to defaults
            from openpyxl.styles import Font, PatternFill, Border, Alignment
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.number_format = "General"
            cell.alignment = Alignment()

        cleared += 1

    mode = "values+formatting" if clear_all else "values"
    return OpResult(success=True, message=f"Cleared {cleared} cells ({mode}) in {range_str}", prefix="-")


def _parse_cell_value(s: str) -> str | int | float:
    """Parse a raw value string into the appropriate Python type.

    Rules:
    - Starts with '=' → formula (string)
    - Wrapped in quotes → text (strip quotes)
    - Leading zero with length > 1 → text (preserve leading zeros, C1)
    - Valid int → int
    - Valid float → float
    - Everything else → text
    """
    # Formula
    if s.startswith("="):
        return s

    # Quoted string — strip outer quotes
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        return s[1:-1]

    # Leading zero protection (C1)
    if len(s) > 1 and s[0] == "0" and s[1:].isdigit():
        return s

    # Try int
    try:
        return int(s)
    except ValueError:
        pass

    # Try float
    try:
        return float(s)
    except ValueError:
        pass

    return s


HANDLERS: dict[str, callable] = {
    "set": op_set,
    "fill": op_fill,
    "clear": op_clear,
    # "data" is handled by adapter block mode, not dispatched here
}
