"""Structure operation handlers — merge, freeze, filter, width, height, etc.

Implements worksheet structure verbs: cell merging, pane freezing,
auto-filter, column width, row height, hide/unhide, group/ungroup.
"""

from __future__ import annotations

from openpyxl.styles import Alignment

from fcp_core import OpResult, ParsedOp

from fcp_sheets.model.refs import (
    ColRef,
    RangeRef,
    RowRef,
    col_to_index,
    index_to_col,
    parse_cell_ref,
    parse_range_ref,
)
from fcp_sheets.server.resolvers import SheetsOpContext


def op_merge(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Merge cells in a range.

    Syntax: merge RANGE [align:center]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: merge RANGE [align:center]")

    range_str = op.positionals[0]
    ws = ctx.active_sheet

    # Validate range
    rr = parse_range_ref(range_str)
    if not isinstance(rr, RangeRef):
        return OpResult(success=False, message=f"Invalid range for merge: {range_str!r}")

    ws.merge_cells(range_str)

    # Optional alignment
    align = op.params.get("align")
    if align:
        # Apply alignment to the top-left cell of the merged region
        cell = ws.cell(row=rr.start.row, column=rr.start.col)
        cell.alignment = Alignment(horizontal=align)

    return OpResult(
        success=True,
        message=f"Merged {range_str}",
        prefix="*",
    )


def op_unmerge(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Unmerge a previously merged range.

    Syntax: unmerge RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: unmerge RANGE")

    range_str = op.positionals[0]
    ws = ctx.active_sheet

    ws.unmerge_cells(range_str)

    return OpResult(
        success=True,
        message=f"Unmerged {range_str}",
        prefix="*",
    )


def op_freeze(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Freeze panes at a cell position.

    Syntax: freeze CELL
    e.g. freeze A2 (freeze top row), freeze B1 (freeze first column)
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: freeze CELL")

    cell_str = op.positionals[0]
    ref = parse_cell_ref(cell_str)
    if not ref:
        return OpResult(success=False, message=f"Invalid cell reference: {cell_str!r}")

    ws = ctx.active_sheet
    ws.freeze_panes = cell_str.upper()

    return OpResult(
        success=True,
        message=f"Frozen panes at {cell_str.upper()}",
        prefix="*",
    )


def op_unfreeze(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Remove pane freeze.

    Syntax: unfreeze
    """
    ws = ctx.active_sheet
    ws.freeze_panes = None

    return OpResult(
        success=True,
        message="Panes unfrozen",
        prefix="*",
    )


def op_filter(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Set or remove auto-filter.

    Syntax: filter RANGE | filter off
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: filter RANGE | filter off")

    ws = ctx.active_sheet

    if op.positionals[0].lower() == "off":
        ws.auto_filter.ref = None
        return OpResult(success=True, message="Auto-filter removed", prefix="*")

    range_str = op.positionals[0]
    rr = parse_range_ref(range_str)
    if not isinstance(rr, RangeRef):
        return OpResult(success=False, message=f"Invalid range for filter: {range_str!r}")

    ws.auto_filter.ref = range_str

    return OpResult(
        success=True,
        message=f"Auto-filter set on {range_str}",
        prefix="*",
    )


def _parse_col_spec(spec: str) -> list[str]:
    """Parse a column specifier into a list of column letters.

    Handles:
      "A"     -> ["A"]
      "A:E"   -> ["A", "B", "C", "D", "E"]
    """
    spec = spec.strip().upper()
    if ":" in spec:
        parts = spec.split(":", 1)
        start = col_to_index(parts[0])
        end = col_to_index(parts[1])
        return [index_to_col(i) for i in range(start, end + 1)]
    return [spec]


def _parse_row_spec(spec: str) -> list[int]:
    """Parse a row specifier into a list of row numbers.

    Handles:
      "3"     -> [3]
      "3:7"   -> [3, 4, 5, 6, 7]
    """
    spec = spec.strip()
    if ":" in spec:
        parts = spec.split(":", 1)
        start = int(parts[0])
        end = int(parts[1])
        return list(range(start, end + 1))
    return [int(spec)]


def op_width(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Set column width.

    Syntax: width COL|RANGE SIZE|auto
    e.g. width A 12, width A:E 14, width A auto
    """
    if len(op.positionals) < 2:
        return OpResult(success=False, message="Usage: width COL|RANGE SIZE|auto")

    col_spec = op.positionals[0]
    size_str = op.positionals[1]
    ws = ctx.active_sheet

    cols = _parse_col_spec(col_spec)

    if size_str.lower() == "auto":
        # Auto-width: scan cells and compute
        for col_letter in cols:
            col_idx = col_to_index(col_letter)
            max_len = 0
            bounds = ctx.index.get_bounds(ws.title)
            if bounds:
                min_row, _, max_row, _ = bounds
                for row in range(min_row, max_row + 1):
                    val = ws.cell(row=row, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
            width = max_len * 1.2 + 2
            width = max(width, 8)  # minimum width
            ws.column_dimensions[col_letter].width = width

        return OpResult(
            success=True,
            message=f"Auto-sized column(s) {col_spec}",
            prefix="*",
        )

    try:
        size = float(size_str)
    except ValueError:
        return OpResult(success=False, message=f"Invalid size: {size_str!r}")

    for col_letter in cols:
        ws.column_dimensions[col_letter].width = size

    return OpResult(
        success=True,
        message=f"Set width of {col_spec} to {size}",
        prefix="*",
    )


def op_height(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Set row height.

    Syntax: height ROW|RANGE SIZE
    e.g. height 1 30, height 1:5 25
    """
    if len(op.positionals) < 2:
        return OpResult(success=False, message="Usage: height ROW|RANGE SIZE")

    row_spec = op.positionals[0]
    size_str = op.positionals[1]

    try:
        size = float(size_str)
    except ValueError:
        return OpResult(success=False, message=f"Invalid size: {size_str!r}")

    ws = ctx.active_sheet
    rows = _parse_row_spec(row_spec)

    for row in rows:
        ws.row_dimensions[row].height = size

    return OpResult(
        success=True,
        message=f"Set height of row(s) {row_spec} to {size}",
        prefix="*",
    )


def op_hide_col(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Hide column(s).

    Syntax: hide-col COL|RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: hide-col COL|RANGE")

    col_spec = op.positionals[0]
    ws = ctx.active_sheet
    cols = _parse_col_spec(col_spec)

    for col_letter in cols:
        ws.column_dimensions[col_letter].hidden = True

    return OpResult(
        success=True,
        message=f"Hidden column(s) {col_spec}",
        prefix="*",
    )


def op_hide_row(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Hide row(s).

    Syntax: hide-row ROW|RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: hide-row ROW|RANGE")

    row_spec = op.positionals[0]
    ws = ctx.active_sheet
    rows = _parse_row_spec(row_spec)

    for row in rows:
        ws.row_dimensions[row].hidden = True

    return OpResult(
        success=True,
        message=f"Hidden row(s) {row_spec}",
        prefix="*",
    )


def op_unhide_col(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Unhide column(s).

    Syntax: unhide-col COL|RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: unhide-col COL|RANGE")

    col_spec = op.positionals[0]
    ws = ctx.active_sheet
    cols = _parse_col_spec(col_spec)

    for col_letter in cols:
        ws.column_dimensions[col_letter].hidden = False

    return OpResult(
        success=True,
        message=f"Unhidden column(s) {col_spec}",
        prefix="*",
    )


def op_unhide_row(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Unhide row(s).

    Syntax: unhide-row ROW|RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: unhide-row ROW|RANGE")

    row_spec = op.positionals[0]
    ws = ctx.active_sheet
    rows = _parse_row_spec(row_spec)

    for row in rows:
        ws.row_dimensions[row].hidden = False

    return OpResult(
        success=True,
        message=f"Unhidden row(s) {row_spec}",
        prefix="*",
    )


def op_group_rows(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Group rows with optional collapse.

    Syntax: group-rows RANGE [collapse]
    e.g. group-rows 2:5, group-rows 2:5 collapse
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: group-rows ROW_RANGE [collapse]")

    row_spec = op.positionals[0]
    collapse = "collapse" in [p.lower() for p in op.positionals[1:]]

    rows = _parse_row_spec(row_spec)
    if len(rows) < 1:
        return OpResult(success=False, message=f"Invalid row range: {row_spec!r}")

    start_row = min(rows)
    end_row = max(rows)
    ws = ctx.active_sheet

    ws.row_dimensions.group(start_row, end_row, outline_level=1, hidden=collapse)

    msg = f"Grouped rows {start_row}:{end_row}"
    if collapse:
        msg += " (collapsed)"
    return OpResult(success=True, message=msg, prefix="*")


def op_group_cols(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Group columns with optional collapse.

    Syntax: group-cols RANGE [collapse]
    e.g. group-cols B:D, group-cols B:D collapse
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: group-cols COL_RANGE [collapse]")

    col_spec = op.positionals[0]
    collapse = "collapse" in [p.lower() for p in op.positionals[1:]]

    cols = _parse_col_spec(col_spec)
    if len(cols) < 1:
        return OpResult(success=False, message=f"Invalid column range: {col_spec!r}")

    start_col = cols[0]
    end_col = cols[-1]
    ws = ctx.active_sheet

    ws.column_dimensions.group(start_col, end_col, outline_level=1, hidden=collapse)

    msg = f"Grouped columns {start_col}:{end_col}"
    if collapse:
        msg += " (collapsed)"
    return OpResult(success=True, message=msg, prefix="*")


def op_ungroup_rows(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Ungroup rows.

    Syntax: ungroup-rows RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: ungroup-rows ROW_RANGE")

    row_spec = op.positionals[0]
    rows = _parse_row_spec(row_spec)
    start_row = min(rows)
    end_row = max(rows)
    ws = ctx.active_sheet

    # Remove grouping by setting outline_level to 0
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].outlineLevel = 0
        ws.row_dimensions[row].hidden = False

    return OpResult(
        success=True,
        message=f"Ungrouped rows {start_row}:{end_row}",
        prefix="*",
    )


def op_ungroup_cols(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Ungroup columns.

    Syntax: ungroup-cols RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: ungroup-cols COL_RANGE")

    col_spec = op.positionals[0]
    cols = _parse_col_spec(col_spec)
    ws = ctx.active_sheet

    for col_letter in cols:
        ws.column_dimensions[col_letter].outlineLevel = 0
        ws.column_dimensions[col_letter].hidden = False

    return OpResult(
        success=True,
        message=f"Ungrouped columns {cols[0]}:{cols[-1]}",
        prefix="*",
    )


HANDLERS: dict[str, callable] = {
    "merge": op_merge,
    "unmerge": op_unmerge,
    "freeze": op_freeze,
    "unfreeze": op_unfreeze,
    "filter": op_filter,
    "width": op_width,
    "height": op_height,
    "hide-col": op_hide_col,
    "hide-row": op_hide_row,
    "unhide-col": op_unhide_col,
    "unhide-row": op_unhide_row,
    "group-rows": op_group_rows,
    "group-cols": op_group_cols,
    "ungroup-rows": op_ungroup_rows,
    "ungroup-cols": op_ungroup_cols,
}
