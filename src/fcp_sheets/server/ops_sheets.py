"""Sheet management handlers — add, remove, rename, copy, hide, unhide, activate."""

from __future__ import annotations

from copy import copy

from openpyxl import Workbook

from fcp_core import OpResult, ParsedOp

from fcp_sheets.server.resolvers import SheetsOpContext


def op_sheet(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Dispatch sheet sub-commands."""
    if not op.positionals:
        return OpResult(success=False, message="Usage: sheet add|remove|rename|copy|hide|unhide|activate NAME")

    action = op.positionals[0].lower()
    rest = op.positionals[1:]

    dispatch = {
        "add": _sheet_add,
        "remove": _sheet_remove,
        "rename": _sheet_rename,
        "copy": _sheet_copy,
        "hide": _sheet_hide,
        "unhide": _sheet_unhide,
        "activate": _sheet_activate,
    }

    handler = dispatch.get(action)
    if handler is None:
        return OpResult(
            success=False,
            message=f"Unknown sheet action: {action!r}. Use: add, remove, rename, copy, hide, unhide, activate",
        )

    return handler(rest, op.params, ctx)


def _sheet_add(
    args: list[str], params: dict[str, str], ctx: SheetsOpContext
) -> OpResult:
    """Add a new worksheet."""
    if not args:
        return OpResult(success=False, message="Usage: sheet add NAME [at:N]")

    name = args[0]
    if name in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{name}' already exists")

    at_index = None
    if "at" in params:
        try:
            at_index = int(params["at"])
        except ValueError:
            return OpResult(success=False, message=f"Invalid position: {params['at']!r}")

    ws = ctx.wb.create_sheet(title=name, index=at_index)
    ctx.index.active_sheet = ws.title
    ctx.wb.active = ctx.wb.sheetnames.index(ws.title)
    return OpResult(success=True, message=f"Sheet '{name}' added", prefix="+")


def _sheet_remove(
    args: list[str], params: dict[str, str], ctx: SheetsOpContext
) -> OpResult:
    """Remove a worksheet."""
    if not args:
        return OpResult(success=False, message="Usage: sheet remove NAME")

    name = args[0]
    if name not in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{name}' not found")

    if len(ctx.wb.sheetnames) <= 1:
        return OpResult(success=False, message="Cannot remove the last sheet")

    was_active = ctx.index.active_sheet == name

    del ctx.wb[name]
    ctx.index.remove_sheet(name)

    # If the removed sheet was active, pick a valid remaining sheet
    if was_active:
        remaining = ctx.wb.sheetnames
        if remaining:
            ctx.index.active_sheet = remaining[0]
            ctx.wb.active = 0
        else:
            ctx.index.active_sheet = ""
    return OpResult(success=True, message=f"Sheet '{name}' removed", prefix="-")


def _sheet_rename(
    args: list[str], params: dict[str, str], ctx: SheetsOpContext
) -> OpResult:
    """Rename a worksheet."""
    if len(args) < 2:
        return OpResult(success=False, message='Usage: sheet rename OLD_NAME "New Name"')

    old_name = args[0]
    new_name = args[1]

    if old_name not in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{old_name}' not found")

    if new_name in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{new_name}' already exists")

    ctx.wb[old_name].title = new_name
    ctx.index.rename_sheet(old_name, new_name)
    return OpResult(success=True, message=f"Sheet '{old_name}' → '{new_name}'", prefix="*")


def _sheet_copy(
    args: list[str], params: dict[str, str], ctx: SheetsOpContext
) -> OpResult:
    """Copy a worksheet."""
    if len(args) < 2:
        return OpResult(success=False, message='Usage: sheet copy SOURCE "Copy Name"')

    source_name = args[0]
    copy_name = args[1]

    if source_name not in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{source_name}' not found")

    if copy_name in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{copy_name}' already exists")

    source_ws = ctx.wb[source_name]
    new_ws = ctx.wb.copy_worksheet(source_ws)
    new_ws.title = copy_name
    return OpResult(success=True, message=f"Sheet '{source_name}' copied as '{copy_name}'", prefix="+")


def _sheet_hide(
    args: list[str], params: dict[str, str], ctx: SheetsOpContext
) -> OpResult:
    """Hide a worksheet."""
    if not args:
        return OpResult(success=False, message="Usage: sheet hide NAME")

    name = args[0]
    if name not in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{name}' not found")

    ctx.wb[name].sheet_state = "hidden"
    return OpResult(success=True, message=f"Sheet '{name}' hidden", prefix="*")


def _sheet_unhide(
    args: list[str], params: dict[str, str], ctx: SheetsOpContext
) -> OpResult:
    """Unhide a worksheet."""
    if not args:
        return OpResult(success=False, message="Usage: sheet unhide NAME")

    name = args[0]
    if name not in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{name}' not found")

    ctx.wb[name].sheet_state = "visible"
    return OpResult(success=True, message=f"Sheet '{name}' unhidden", prefix="*")


def _sheet_activate(
    args: list[str], params: dict[str, str], ctx: SheetsOpContext
) -> OpResult:
    """Switch active worksheet."""
    if not args:
        return OpResult(success=False, message="Usage: sheet activate NAME")

    name = args[0]
    if name not in ctx.wb.sheetnames:
        return OpResult(success=False, message=f"Sheet '{name}' not found")

    ctx.wb.active = ctx.wb.sheetnames.index(name)
    ctx.index.active_sheet = name
    return OpResult(success=True, message=f"Active sheet: '{name}'", prefix="*")


HANDLERS: dict[str, callable] = {
    "sheet": op_sheet,
}
