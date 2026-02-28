"""Conditional formatting handlers — cond-fmt variants.

Supports: color-scale, data-bar, icon-set, cell-is, formula, duplicate, unique, top, bottom.
"""

from __future__ import annotations

from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)
from openpyxl.styles import Font, PatternFill

from fcp_core import OpResult, ParsedOp

from fcp_sheets.lib.colors import parse_color
from fcp_sheets.server.resolvers import SheetsOpContext

# Operator mapping from DSL names to openpyxl operator names
_OP_MAP: dict[str, str] = {
    "gt": "greaterThan",
    "lt": "lessThan",
    "gte": "greaterThanOrEqual",
    "lte": "lessThanOrEqual",
    "eq": "equal",
    "neq": "notEqual",
    "ne": "notEqual",
    "between": "between",
    "not-between": "notBetween",
}

# Icon set style mapping
_ICON_MAP: dict[str, str] = {
    "arrows": "3Arrows",
    "flags": "3Flags",
    "traffic": "3TrafficLights1",
    "rating": "3Stars",
    "symbols": "3Symbols",
    "3arrows": "3Arrows",
    "3flags": "3Flags",
    "3traffic": "3TrafficLights1",
    "3stars": "3Stars",
    "3symbols": "3Symbols",
    "4arrows": "4Arrows",
    "4traffic": "4TrafficLights",
    "5arrows": "5Arrows",
    "5rating": "5Rating",
}


def _make_fill(color_hex: str | None) -> PatternFill | None:
    """Create a PatternFill from a hex color string, or None."""
    if color_hex is None:
        return None
    try:
        c = parse_color(color_hex)
        return PatternFill(start_color=c, end_color=c, fill_type="solid")
    except ValueError:
        return None


def _make_font(color_hex: str | None = None, bold: bool = False) -> Font | None:
    """Create a Font from optional color and bold flag, or None if both default."""
    if color_hex is None and not bold:
        return None
    font_color = None
    if color_hex:
        try:
            font_color = parse_color(color_hex)
        except ValueError:
            pass
    return Font(bold=bold, color=font_color)


def _cond_fmt_color_scale(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE color-scale [min-color:#HEX] [max-color:#HEX] [mid-color:#HEX]"""
    ws = ctx.active_sheet

    min_color = op.params.get("min-color", "#F8696B")
    max_color = op.params.get("max-color", "#63BE7B")
    mid_color = op.params.get("mid-color")

    try:
        min_c = parse_color(min_color)
        max_c = parse_color(max_color)
    except ValueError as e:
        return OpResult(success=False, message=str(e))

    if mid_color:
        try:
            mid_c = parse_color(mid_color)
        except ValueError as e:
            return OpResult(success=False, message=str(e))
        rule = ColorScaleRule(
            start_type="min", start_color=min_c,
            mid_type="percentile", mid_value=50, mid_color=mid_c,
            end_type="max", end_color=max_c,
        )
    else:
        rule = ColorScaleRule(
            start_type="min", start_color=min_c,
            end_type="max", end_color=max_c,
        )

    ws.conditional_formatting.add(range_str, rule)
    colors = "3-color" if mid_color else "2-color"
    return OpResult(success=True, message=f"Color scale ({colors}) on {range_str}", prefix="+")


def _cond_fmt_data_bar(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE data-bar [color:#HEX]"""
    ws = ctx.active_sheet

    color = op.params.get("color", "#638EC6")
    try:
        c = parse_color(color)
    except ValueError as e:
        return OpResult(success=False, message=str(e))

    rule = DataBarRule(start_type="min", end_type="max", color=c)
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Data bar on {range_str}", prefix="+")


def _cond_fmt_icon_set(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE icon-set [icons:arrows|flags|traffic|rating|symbols]"""
    ws = ctx.active_sheet

    icons_name = op.params.get("icons", "arrows").lower()
    icon_style = _ICON_MAP.get(icons_name)
    if icon_style is None:
        available = ", ".join(sorted(_ICON_MAP.keys()))
        return OpResult(success=False, message=f"Unknown icon set: {icons_name!r}. Available: {available}")

    rule = IconSetRule(icon_style=icon_style, type="percent", values=[0, 33, 67])
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Icon set ({icons_name}) on {range_str}", prefix="+")


def _cond_fmt_cell_is(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE cell-is OP VALUE [VALUE2] [fill:#HEX] [color:#HEX] [bold]"""
    ws = ctx.active_sheet

    # Expect: positionals = [RANGE, "cell-is", OP, VALUE, ...]
    if len(op.positionals) < 4:
        return OpResult(
            success=False,
            message="Usage: cond-fmt RANGE cell-is OP VALUE [VALUE2] [fill:#HEX] [color:#HEX] [bold]",
        )

    op_name = op.positionals[2].lower()
    mapped_op = _OP_MAP.get(op_name)
    if mapped_op is None:
        available = ", ".join(sorted(_OP_MAP.keys()))
        return OpResult(success=False, message=f"Unknown operator: {op_name!r}. Available: {available}")

    value1 = op.positionals[3]
    formula = [str(value1)]

    # For between/not-between, need a second value
    if mapped_op in ("between", "notBetween"):
        if len(op.positionals) < 5:
            return OpResult(success=False, message=f"Operator '{op_name}' requires two values")
        value2 = op.positionals[4]
        formula.append(str(value2))

    # Build formatting
    fill_color = op.params.get("fill")
    font_color = op.params.get("color")
    bold_flag = "bold" in [p.lower() for p in op.positionals]

    fill = _make_fill(fill_color)
    font = _make_font(font_color, bold_flag)

    kwargs: dict = {"operator": mapped_op, "formula": formula}
    if fill:
        kwargs["fill"] = fill
    if font:
        kwargs["font"] = font

    rule = CellIsRule(**kwargs)
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Cell-is ({op_name}) on {range_str}", prefix="+")


def _cond_fmt_formula(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE formula FORMULA [fill:#HEX] [color:#HEX] [bold]"""
    ws = ctx.active_sheet

    if len(op.positionals) < 3:
        return OpResult(
            success=False,
            message="Usage: cond-fmt RANGE formula FORMULA [fill:#HEX] [color:#HEX] [bold]",
        )

    formula_str = op.positionals[2]

    fill_color = op.params.get("fill")
    font_color = op.params.get("color")
    bold_flag = "bold" in [p.lower() for p in op.positionals]

    fill = _make_fill(fill_color)
    font = _make_font(font_color, bold_flag)

    kwargs: dict = {"formula": [formula_str]}
    if fill:
        kwargs["fill"] = fill
    if font:
        kwargs["font"] = font

    rule = FormulaRule(**kwargs)
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Formula rule on {range_str}", prefix="+")


def _cond_fmt_duplicate(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE duplicate [fill:#HEX]"""
    ws = ctx.active_sheet

    fill_color = op.params.get("fill", "#FFC7CE")
    fill = _make_fill(fill_color)

    # COUNTIF formula to detect duplicates
    # Use the first cell of the range as the reference
    first_cell = range_str.split(":")[0]
    formula_str = f"COUNTIF({range_str},{first_cell})>1"

    kwargs: dict = {"formula": [formula_str]}
    if fill:
        kwargs["fill"] = fill

    rule = FormulaRule(**kwargs)
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Duplicate highlighting on {range_str}", prefix="+")


def _cond_fmt_unique(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE unique [fill:#HEX]"""
    ws = ctx.active_sheet

    fill_color = op.params.get("fill", "#C6EFCE")
    fill = _make_fill(fill_color)

    first_cell = range_str.split(":")[0]
    formula_str = f"COUNTIF({range_str},{first_cell})=1"

    kwargs: dict = {"formula": [formula_str]}
    if fill:
        kwargs["fill"] = fill

    rule = FormulaRule(**kwargs)
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Unique highlighting on {range_str}", prefix="+")


def _cond_fmt_top(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE top N [fill:#HEX]"""
    ws = ctx.active_sheet

    if len(op.positionals) < 3:
        return OpResult(success=False, message="Usage: cond-fmt RANGE top N [fill:#HEX]")

    try:
        n = int(op.positionals[2])
    except ValueError:
        return OpResult(success=False, message=f"Invalid count: {op.positionals[2]!r}")

    fill_color = op.params.get("fill", "#C6EFCE")
    fill = _make_fill(fill_color)

    first_cell = range_str.split(":")[0]
    formula_str = f"RANK({first_cell},{range_str})<={n}"

    kwargs: dict = {"formula": [formula_str]}
    if fill:
        kwargs["fill"] = fill

    rule = FormulaRule(**kwargs)
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Top {n} on {range_str}", prefix="+")


def _cond_fmt_bottom(op: ParsedOp, ctx: SheetsOpContext, range_str: str) -> OpResult:
    """cond-fmt RANGE bottom N [fill:#HEX]"""
    ws = ctx.active_sheet

    if len(op.positionals) < 3:
        return OpResult(success=False, message="Usage: cond-fmt RANGE bottom N [fill:#HEX]")

    try:
        n = int(op.positionals[2])
    except ValueError:
        return OpResult(success=False, message=f"Invalid count: {op.positionals[2]!r}")

    fill_color = op.params.get("fill", "#FFC7CE")
    fill = _make_fill(fill_color)

    first_cell = range_str.split(":")[0]
    formula_str = f"RANK({first_cell},{range_str})>=COUNT({range_str})-{n}+1"

    kwargs: dict = {"formula": [formula_str]}
    if fill:
        kwargs["fill"] = fill

    rule = FormulaRule(**kwargs)
    ws.conditional_formatting.add(range_str, rule)

    return OpResult(success=True, message=f"Bottom {n} on {range_str}", prefix="+")


# Sub-type dispatch
_COND_FMT_TYPES: dict[str, callable] = {
    "color-scale": _cond_fmt_color_scale,
    "data-bar": _cond_fmt_data_bar,
    "icon-set": _cond_fmt_icon_set,
    "cell-is": _cond_fmt_cell_is,
    "formula": _cond_fmt_formula,
    "duplicate": _cond_fmt_duplicate,
    "unique": _cond_fmt_unique,
    "top": _cond_fmt_top,
    "bottom": _cond_fmt_bottom,
}


def op_cond_fmt(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Main cond-fmt verb dispatcher.

    Syntax: cond-fmt RANGE TYPE [params...]
    """
    if len(op.positionals) < 2:
        return OpResult(
            success=False,
            message="Usage: cond-fmt RANGE TYPE [params...]",
        )

    range_str = op.positionals[0]
    fmt_type = op.positionals[1].lower()

    handler = _COND_FMT_TYPES.get(fmt_type)
    if handler is None:
        available = ", ".join(sorted(_COND_FMT_TYPES.keys()))
        return OpResult(success=False, message=f"Unknown cond-fmt type: {fmt_type!r}. Available: {available}")

    return handler(op, ctx, range_str)


HANDLERS: dict[str, callable] = {
    "cond-fmt": op_cond_fmt,
}
