"""Style operation handlers — style, border, define-style, apply-style.

Implements cell formatting using openpyxl styles (Font, PatternFill,
Alignment, Border, Side) with the copy() pattern to preserve existing
properties when only modifying specific attributes.
"""

from __future__ import annotations

from copy import copy

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from fcp_core import OpResult, ParsedOp

from fcp_sheets.lib.colors import parse_color
from fcp_sheets.lib.number_formats import resolve_format
from fcp_sheets.model.refs import index_to_col
from fcp_sheets.server.resolvers import (
    SheetsOpContext,
    resolve_range_to_cells,
    resolve_target_cells,
)

# Font-related flags that can appear as positionals
_FONT_FLAGS = {"bold", "italic", "underline", "strike"}
# Alignment flag
_ALIGN_FLAGS = {"wrap"}
# All recognized flags
_ALL_FLAGS = _FONT_FLAGS | _ALIGN_FLAGS


def op_style(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Apply formatting to a range of cells.

    Syntax:
      style RANGE|@SEL [bold] [italic] [underline] [strike]
        [font:NAME] [size:N] [color:#HEX]
        [fill:#HEX] [fill-pattern:PATTERN]
        [align:left|center|right] [valign:top|middle|bottom]
        [wrap] [indent:N] [rotate:N]
        [fmt:FORMAT]

    Flags (bold, italic, underline, strike, wrap) are parsed from positionals.
    Key:value params handle the rest.
    """
    if not op.positionals and not op.selectors:
        return OpResult(
            success=False,
            message="Usage: style RANGE|@SEL [bold] [italic] ... [font:NAME] [size:N] ...",
        )

    # Separate the range/selector from flag positionals
    positionals = list(op.positionals)
    flags: set[str] = set()
    range_positional: str | None = None

    for p in positionals:
        low = p.lower()
        if low in _ALL_FLAGS:
            flags.add(low)
        elif range_positional is None:
            range_positional = p
        else:
            # Extra positional — treat as flag if recognized
            if low in _ALL_FLAGS:
                flags.add(low)

    # Resolve target cells
    target_positionals = [range_positional] if range_positional else []
    cells = resolve_target_cells(target_positionals, op.selectors, ctx)
    if not cells:
        return OpResult(success=False, message="No cells resolved for style")

    # Extract params
    params = op.params
    font_name = params.get("font")
    font_size = params.get("size")
    font_color = params.get("color")
    fill_color = params.get("fill")
    fill_pattern = params.get("fill-pattern", "solid")
    h_align = params.get("align")
    v_align = params.get("valign")
    # Map "middle" to "center" for openpyxl compatibility
    if v_align == "middle":
        v_align = "center"
    indent = params.get("indent")
    rotate = params.get("rotate")
    fmt = params.get("fmt")

    count = 0
    for ws, row, col in cells:
        cell = ws.cell(row=row, column=col)

        # -- Font (copy pattern) --
        if flags & _FONT_FLAGS or font_name or font_size or font_color:
            old = copy(cell.font)
            kwargs: dict = {}
            kwargs["name"] = font_name if font_name else old.name
            kwargs["size"] = float(font_size) if font_size else old.size
            if font_color:
                kwargs["color"] = parse_color(font_color)
            elif old.color:
                kwargs["color"] = old.color
            kwargs["bold"] = True if "bold" in flags else old.bold
            kwargs["italic"] = True if "italic" in flags else old.italic
            if "underline" in flags:
                kwargs["underline"] = "single"
            elif old.underline:
                kwargs["underline"] = old.underline
            kwargs["strike"] = True if "strike" in flags else old.strike
            cell.font = Font(**kwargs)

        # -- Fill --
        if fill_color:
            hex_color = parse_color(fill_color)
            cell.fill = PatternFill(
                start_color=hex_color,
                end_color=hex_color,
                fill_type=fill_pattern,
            )

        # -- Alignment --
        if h_align or v_align or "wrap" in flags or indent or rotate:
            old_align = copy(cell.alignment)
            align_kwargs: dict = {}
            align_kwargs["horizontal"] = h_align if h_align else old_align.horizontal
            align_kwargs["vertical"] = v_align if v_align else old_align.vertical
            align_kwargs["wrap_text"] = True if "wrap" in flags else old_align.wrap_text
            if indent:
                align_kwargs["indent"] = int(indent)
            elif old_align.indent:
                align_kwargs["indent"] = old_align.indent
            if rotate:
                align_kwargs["text_rotation"] = int(rotate)
            elif old_align.text_rotation:
                align_kwargs["text_rotation"] = old_align.text_rotation
            cell.alignment = Alignment(**align_kwargs)

        # -- Number format --
        if fmt:
            cell.number_format = resolve_format(fmt)

        count += 1

    # Build summary of what was applied
    parts: list[str] = []
    if flags:
        parts.append(", ".join(sorted(flags)))
    for k in ("font", "size", "color", "fill", "align", "valign", "fmt"):
        if k in params:
            parts.append(f"{k}:{params[k]}")
    desc = "; ".join(parts) if parts else "style"
    return OpResult(
        success=True,
        message=f"Styled {count} cells ({desc})",
        prefix="*",
    )


# Valid border line styles
_LINE_STYLES = {
    "thin", "medium", "thick", "dashed", "dotted", "double",
    "hair", "mediumDashed", "dashDot", "mediumDashDot",
    "dashDotDot", "mediumDashDotDot", "slantDashDot",
}

# Side keywords
_SIDE_KEYWORDS = {
    "all", "outline", "top", "bottom", "left", "right",
    "inner", "h", "v",
}


def op_border(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Apply borders to a range of cells.

    Syntax:
      border RANGE|@SEL SIDES [line:STYLE] [color:#HEX]
        SIDES: all | outline | top | bottom | left | right | inner | h | v
        STYLE: thin | medium | thick | dashed | dotted | double | hair
    """
    if len(op.positionals) < 2 and not op.selectors:
        return OpResult(
            success=False,
            message="Usage: border RANGE SIDES [line:STYLE] [color:#HEX]",
        )

    positionals = list(op.positionals)
    range_positional: str | None = None
    sides_str: str | None = None

    # Parse positionals: first is range, rest are sides keywords
    for p in positionals:
        low = p.lower()
        if low in _SIDE_KEYWORDS:
            sides_str = low
        elif range_positional is None:
            range_positional = p
        elif sides_str is None:
            sides_str = low

    if not sides_str:
        return OpResult(
            success=False,
            message="Missing border sides. Use: all | outline | top | bottom | left | right | inner | h | v",
        )

    # Resolve target cells
    target_positionals = [range_positional] if range_positional else []
    cells = resolve_target_cells(target_positionals, op.selectors, ctx)
    if not cells:
        return OpResult(success=False, message="No cells resolved for border")

    # Get line style and color
    line_style = op.params.get("line", "thin")
    if line_style not in _LINE_STYLES:
        return OpResult(success=False, message=f"Invalid line style: {line_style!r}")

    color_hex = None
    if "color" in op.params:
        color_hex = parse_color(op.params["color"])

    side = Side(style=line_style, color=color_hex)
    no_side = Side(style=None)

    # Determine bounding box for outline/inner
    rows = {r for _, r, _ in cells}
    cols = {c for _, _, c in cells}
    min_row, max_row = min(rows), max(rows)
    min_col, max_col = min(cols), max(cols)

    count = 0
    for ws, row, col in cells:
        cell = ws.cell(row=row, column=col)
        old_border = copy(cell.border)

        top = copy(old_border.top)
        bottom = copy(old_border.bottom)
        left = copy(old_border.left)
        right = copy(old_border.right)

        if sides_str == "all":
            top = bottom = left = right = side
        elif sides_str == "outline":
            if row == min_row:
                top = side
            if row == max_row:
                bottom = side
            if col == min_col:
                left = side
            if col == max_col:
                right = side
        elif sides_str == "top":
            top = side
        elif sides_str == "bottom":
            bottom = side
        elif sides_str == "left":
            left = side
        elif sides_str == "right":
            right = side
        elif sides_str == "inner":
            if row > min_row:
                top = side
            if row < max_row:
                bottom = side
            if col > min_col:
                left = side
            if col < max_col:
                right = side
        elif sides_str == "h":
            # Horizontal internal borders
            if row > min_row:
                top = side
            if row < max_row:
                bottom = side
        elif sides_str == "v":
            # Vertical internal borders
            if col > min_col:
                left = side
            if col < max_col:
                right = side

        cell.border = Border(top=top, bottom=bottom, left=left, right=right)
        count += 1

    return OpResult(
        success=True,
        message=f"Applied {sides_str} border ({line_style}) to {count} cells",
        prefix="*",
    )


def op_define_style(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Define a named style for reuse.

    Syntax:
      define-style NAME [font:F] [size:N] [bold] [fill:#HEX] [color:#HEX] [fmt:FORMAT]
    """
    if not op.positionals:
        return OpResult(
            success=False,
            message="Usage: define-style NAME [font:F] [size:N] [bold] [fill:#HEX] ...",
        )

    name = op.positionals[0]

    # Collect style definition from remaining positionals and params
    style_def: dict[str, str | bool] = {}

    # Flags from positionals
    for p in op.positionals[1:]:
        low = p.lower()
        if low in _ALL_FLAGS:
            style_def[low] = True

    # Key:value params
    for k, v in op.params.items():
        style_def[k] = v

    ctx.named_styles[name] = style_def

    return OpResult(
        success=True,
        message=f"Defined style {name!r}",
        prefix="+",
    )


def op_apply_style(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Apply a previously defined named style to a range.

    Syntax:
      apply-style NAME RANGE|@SEL
    """
    if not op.positionals:
        return OpResult(
            success=False,
            message="Usage: apply-style NAME RANGE|@SEL",
        )

    name = op.positionals[0]
    if name not in ctx.named_styles:
        return OpResult(success=False, message=f"Unknown style: {name!r}")

    style_def = ctx.named_styles[name]

    # Build a synthetic ParsedOp for style verb
    # Remaining positionals are the range
    range_positionals = op.positionals[1:]

    # Convert style_def back to params and positionals for op_style
    synth_params: dict[str, str] = {}
    synth_positionals: list[str] = list(range_positionals)

    for k, v in style_def.items():
        if k in _ALL_FLAGS and v is True:
            synth_positionals.append(k)
        else:
            synth_params[str(k)] = str(v)

    synth_op = ParsedOp(
        verb="style",
        positionals=synth_positionals,
        params=synth_params,
        selectors=list(op.selectors),
        raw=op.raw,
    )

    result = op_style(synth_op, ctx)
    if result.success:
        return OpResult(
            success=True,
            message=f"Applied style {name!r} — {result.message}",
            prefix="*",
        )
    return result


HANDLERS: dict[str, callable] = {
    "style": op_style,
    "border": op_border,
    "define-style": op_define_style,
    "apply-style": op_apply_style,
}
