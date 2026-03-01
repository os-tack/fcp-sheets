"""Tests for MVP queries — plan/map, stats, status.
Extended Wave 3 tests — describe, peek, list, find.
"""

from __future__ import annotations

import pytest

from fcp_sheets.adapter import SheetsAdapter
from fcp_sheets.model.snapshot import SheetsModel
from fcp_sheets.server.queries import dispatch_query
from fcp_core import EventLog, parse_op


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _setup_revenue_sheet(adapter: SheetsAdapter, model: SheetsModel) -> None:
    """Populate a Revenue-like sheet with headers, data, and formulas.

    NOTE: parse_op treats colon as key:value separator, so we use
    =B2+C2+D2 instead of =SUM(B2:D2) for formulas in tests.
    """
    log = EventLog()
    # Rename Sheet1 to Revenue
    adapter.dispatch_op(parse_op("sheet rename Sheet1 Revenue"), model, log)
    # Headers
    adapter.dispatch_op(parse_op("set A1 Month"), model, log)
    adapter.dispatch_op(parse_op("set B1 Product_A"), model, log)
    adapter.dispatch_op(parse_op("set C1 Product_B"), model, log)
    adapter.dispatch_op(parse_op("set D1 Product_C"), model, log)
    adapter.dispatch_op(parse_op("set E1 Total"), model, log)
    # Row 2
    adapter.dispatch_op(parse_op("set A2 January"), model, log)
    adapter.dispatch_op(parse_op("set B2 15000"), model, log)
    adapter.dispatch_op(parse_op("set C2 22000"), model, log)
    adapter.dispatch_op(parse_op("set D2 8000"), model, log)
    adapter.dispatch_op(parse_op("set E2 =B2+C2+D2"), model, log)
    # Row 3
    adapter.dispatch_op(parse_op("set A3 February"), model, log)
    adapter.dispatch_op(parse_op("set B3 16000"), model, log)
    adapter.dispatch_op(parse_op("set C3 23000"), model, log)
    adapter.dispatch_op(parse_op("set D3 9000"), model, log)
    adapter.dispatch_op(parse_op("set E3 =B3+C3+D3"), model, log)
    # Row 4
    adapter.dispatch_op(parse_op("set A4 March"), model, log)
    adapter.dispatch_op(parse_op("set B4 17000"), model, log)
    adapter.dispatch_op(parse_op("set C4 21000"), model, log)
    adapter.dispatch_op(parse_op("set D4 10000"), model, log)
    adapter.dispatch_op(parse_op("set E4 =B4+C4+D4"), model, log)


def _setup_simple_data(adapter: SheetsAdapter, model: SheetsModel) -> None:
    """Set up a simple 3x3 data block."""
    log = EventLog()
    adapter.dispatch_op(parse_op("set A1 Name"), model, log)
    adapter.dispatch_op(parse_op("set B1 Score"), model, log)
    adapter.dispatch_op(parse_op("set C1 Grade"), model, log)
    adapter.dispatch_op(parse_op("set A2 Alice"), model, log)
    adapter.dispatch_op(parse_op("set B2 95"), model, log)
    adapter.dispatch_op(parse_op("set C2 A"), model, log)
    adapter.dispatch_op(parse_op("set A3 Bob"), model, log)
    adapter.dispatch_op(parse_op("set B3 82"), model, log)
    adapter.dispatch_op(parse_op("set C3 B"), model, log)


# ===========================================================================
# MVP Queries (existing tests)
# ===========================================================================

class TestPlanQuery:
    def test_empty_workbook(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("plan", model, adapter.index)
        assert "Test Workbook" in result
        assert "1 sheets" in result
        assert "Sheet1" in result
        assert "[active]" in result

    def test_with_data(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Name"), model, log)
        adapter.dispatch_op(parse_op("set B1 Score"), model, log)
        adapter.dispatch_op(parse_op("set A2 Alice"), model, log)
        adapter.dispatch_op(parse_op("set B2 95"), model, log)

        result = dispatch_query("plan", model, adapter.index)
        assert "data:" in result
        assert "A1:" in result  # Data bounds
        assert "next-empty:" in result

    def test_map_alias(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("map", model, adapter.index)
        assert "Test Workbook" in result

    def test_multiple_sheets(self, adapter: SheetsAdapter):
        model = adapter.create_empty("Multi", {"sheets": "3"})
        result = dispatch_query("plan", model, adapter.index)
        assert "3 sheets" in result
        assert "Sheet1" in result
        assert "Sheet2" in result
        assert "Sheet3" in result


class TestStatsQuery:
    def test_empty(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("stats", model, adapter.index)
        assert "Test Workbook" in result
        assert "Data cells: 0" in result

    def test_with_data(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 42"), model, log)
        adapter.dispatch_op(parse_op("set A2 =A1*2"), model, log)

        result = dispatch_query("stats", model, adapter.index)
        assert "Data cells: 1" in result
        assert "Formula cells: 1" in result


class TestStatusQuery:
    def test_status(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("status", model, adapter.index)
        assert "Test Workbook" in result
        assert "unsaved" in result
        assert "Sheet1" in result


class TestUnknownQuery:
    def test_unknown(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("foobar", model, adapter.index)
        assert "Unknown query" in result
        assert "try:" in result


# ===========================================================================
# Wave 3: describe
# ===========================================================================

class TestDescribeSheet:
    def test_describe_empty_sheet(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("describe Sheet1", model, adapter.index)
        assert "Sheet: Sheet1" in result
        assert "(empty)" in result

    def test_describe_sheet_with_data(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        result = dispatch_query("describe Revenue", model, adapter.index)
        assert "Sheet: Revenue" in result
        assert "data:" in result
        assert "columns:" in result
        assert "Month" in result
        assert "Product_A" in result
        assert "sample" in result

    def test_describe_sheet_columns_analysis(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        result = dispatch_query("describe Revenue", model, adapter.index)
        # Should show column types
        assert "text" in result
        assert "number" in result

    def test_describe_sheet_formula_column(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        result = dispatch_query("describe Revenue", model, adapter.index)
        assert "formula" in result
        assert "E" in result

    def test_describe_sheet_sample_data(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        result = dispatch_query("describe Revenue", model, adapter.index)
        assert "January" in result
        assert "15000" in result

    def test_describe_sheet_merged_none(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("describe Sheet1", model, adapter.index)
        assert "merged: (none)" in result

    def test_describe_sheet_with_merged(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        # Use openpyxl directly since parse_op treats colon as key:value
        model.wb.active.merge_cells("A1:C1")
        result = dispatch_query("describe Sheet1", model, adapter.index)
        assert "merged:" in result
        assert "A1:C1" in result

    def test_describe_sheet_cond_fmt_none(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("describe Sheet1", model, adapter.index)
        assert "cond-fmt: (none)" in result

    def test_describe_sheet_with_table(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        _setup_simple_data(adapter, model)
        adapter.dispatch_op(parse_op("table add Scores range:A1:C3"), model, log)
        result = dispatch_query("describe Sheet1", model, adapter.index)
        assert "Scores" in result
        assert "A1:C3" in result

    def test_describe_sheet_tables_none(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("describe Sheet1", model, adapter.index)
        assert "tables: (none)" in result

    def test_describe_no_args(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("describe", model, adapter.index)
        assert "Usage" in result

    def test_describe_invalid_target(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("describe nonexistent", model, adapter.index)
        assert "Cannot resolve" in result

    def test_describe_case_insensitive_sheet(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("describe sheet1", model, adapter.index)
        assert "Sheet: Sheet1" in result


class TestDescribeCell:
    def test_describe_cell_with_text(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "Cell A1" in result
        assert "Hello" in result
        assert "text" in result

    def test_describe_cell_with_number(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 42"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "Cell A1" in result
        assert "42" in result
        assert "number" in result

    def test_describe_cell_with_formula(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 10"), model, log)
        adapter.dispatch_op(parse_op("set A2 =A1*2"), model, log)
        result = dispatch_query("describe A2", model, adapter.index)
        assert "Cell A2" in result
        assert "=A1*2" in result
        assert "formula" in result

    def test_describe_cell_empty(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 hi"), model, log)
        result = dispatch_query("describe B1", model, adapter.index)
        assert "Cell B1" in result
        assert "empty" in result

    def test_describe_cell_font_info(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "font:" in result

    def test_describe_cell_fill_info(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "fill:" in result

    def test_describe_cell_alignment_info(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "alignment:" in result

    def test_describe_cell_border_info(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "border:" in result

    def test_describe_cell_merged_no(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "merged: no" in result

    def test_describe_cell_in_table(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        _setup_simple_data(adapter, model)
        adapter.dispatch_op(parse_op("table add MyTable range:A1:C3"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "table MyTable" in result

    def test_describe_cell_styled(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        adapter.dispatch_op(parse_op("style A1 bold fill:#FF0000"), model, log)
        result = dispatch_query("describe A1", model, adapter.index)
        assert "bold" in result
        assert "FF0000" in result


class TestDescribeRange:
    def test_describe_range_basic(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("describe A1:C3", model, adapter.index)
        assert "Range A1:C3" in result
        assert "size:" in result
        assert "3 rows" in result
        assert "3 cols" in result

    def test_describe_range_types(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("describe A1:C3", model, adapter.index)
        assert "types:" in result
        assert "text" in result
        assert "number" in result

    def test_describe_range_number_range(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("describe B1:B3", model, adapter.index)
        assert "number-range:" in result
        assert "82" in result
        assert "95" in result


# ===========================================================================
# Wave 3: peek
# ===========================================================================

class TestPeekQuery:
    def test_peek_no_args(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("peek", model, adapter.index)
        assert "Usage" in result

    def test_peek_invalid_range(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("peek !!!invalid", model, adapter.index)
        assert "Invalid range" in result or "!" in result

    def test_peek_narrow_table(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("peek A1:C3", model, adapter.index)
        assert "A1:C3" in result
        assert "Name" in result
        assert "Alice" in result
        assert "95" in result
        assert "Bob" in result
        assert "82" in result

    def test_peek_single_cell(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("peek A1", model, adapter.index)
        assert "Hello" in result

    def test_peek_formulas_shown(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        result = dispatch_query("peek E1:E4", model, adapter.index)
        assert "=B" in result  # formula starts with =B2+C2+D2
        assert "Total" in result

    def test_peek_empty_cells(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        adapter.dispatch_op(parse_op("set C1 World"), model, log)
        result = dispatch_query("peek A1:C1", model, adapter.index)
        assert "Hello" in result
        assert "World" in result

    def test_peek_row_cap(self, adapter: SheetsAdapter, model: SheetsModel):
        """Verify that peek caps at 50 rows."""
        log = EventLog()
        # Create 60 rows of data
        for i in range(1, 61):
            adapter.dispatch_op(parse_op(f"set A{i} row{i}"), model, log)
        result = dispatch_query("peek A1:A60", model, adapter.index)
        assert "more rows" in result
        assert "capped at 50" in result

    def test_peek_wide_mode(self, adapter: SheetsAdapter, model: SheetsModel):
        """Verify that peek uses wide mode for 12+ columns."""
        log = EventLog()
        # Create 13 columns
        for i in range(1, 14):
            from fcp_sheets.model.refs import index_to_col
            col_letter = index_to_col(i)
            adapter.dispatch_op(parse_op(f"set {col_letter}1 Header{i}"), model, log)
            adapter.dispatch_op(parse_op(f"set {col_letter}2 Val{i}"), model, log)
        result = dispatch_query("peek A1:M2", model, adapter.index)
        assert "Row" in result
        # In wide mode, should show "more cols" if >10
        assert "more cols" in result

    def test_peek_narrow_column_count(self, adapter: SheetsAdapter, model: SheetsModel):
        """Verify narrow mode for < 12 cols uses pipe-delimited format."""
        _setup_simple_data(adapter, model)
        result = dispatch_query("peek A1:C3", model, adapter.index)
        # Narrow mode uses | delimiters
        assert "|" in result

    def test_peek_sheet_name(self, adapter: SheetsAdapter, model: SheetsModel):
        """Verify that peek shows sheet name."""
        _setup_simple_data(adapter, model)
        result = dispatch_query("peek A1:C3", model, adapter.index)
        assert "Sheet1" in result


# ===========================================================================
# Wave 3: list
# ===========================================================================

class TestListSheets:
    def test_list_sheets_empty(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list sheets", model, adapter.index)
        assert "Sheets" in result
        assert "Sheet1" in result
        assert "(empty)" in result

    def test_list_sheets_with_data(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("list sheets", model, adapter.index)
        assert "Sheet1" in result
        assert "3 rows" in result
        assert "3 cols" in result
        assert "data cells" in result

    def test_list_sheets_multiple(self, adapter: SheetsAdapter):
        model = adapter.create_empty("Multi", {"sheets": "3"})
        result = dispatch_query("list sheets", model, adapter.index)
        assert "Sheet1" in result
        assert "Sheet2" in result
        assert "Sheet3" in result

    def test_list_sheets_active_marker(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list sheets", model, adapter.index)
        assert "[active]" in result

    def test_list_sheets_hidden(self, adapter: SheetsAdapter):
        model = adapter.create_empty("Multi", {"sheets": "2"})
        log = EventLog()
        adapter.dispatch_op(parse_op("sheet hide Sheet2"), model, log)
        result = dispatch_query("list sheets", model, adapter.index)
        assert "[hidden]" in result


class TestListCharts:
    def test_list_charts_none(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list charts", model, adapter.index)
        assert "Charts: (none)" in result

    def test_list_charts_with_chart(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        _setup_simple_data(adapter, model)
        adapter.dispatch_op(
            parse_op('chart add bar title:"Score Chart" data:B1:B3'),
            model, log
        )
        result = dispatch_query("list charts", model, adapter.index)
        assert "Charts" in result
        assert "Score Chart" in result


class TestListFormulas:
    def test_list_formulas_none(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 Hello"), model, log)
        result = dispatch_query("list formulas", model, adapter.index)
        assert "Formulas: (none)" in result

    def test_list_formulas_with_formulas(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        result = dispatch_query("list formulas", model, adapter.index)
        assert "Formulas" in result
        assert "=B" in result  # =B2+C2+D2 pattern
        assert "3 cells" in result

    def test_list_formulas_single(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 10"), model, log)
        adapter.dispatch_op(parse_op("set A2 =A1*2"), model, log)
        result = dispatch_query("list formulas", model, adapter.index)
        assert "=A1*2" in result


class TestListStyles:
    def test_list_styles(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list styles", model, adapter.index)
        # openpyxl always creates a default "Normal" style
        assert "Named styles" in result
        assert "Normal" in result


class TestListNames:
    def test_list_names_none(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list names", model, adapter.index)
        assert "Named ranges: (none)" in result

    def test_list_names_with_defined_name(self, adapter: SheetsAdapter, model: SheetsModel):
        # Add a defined name directly through openpyxl
        from openpyxl.workbook.defined_name import DefinedName
        dn = DefinedName("MyRange", attr_text="Sheet1!$A$1:$C$3")
        model.wb.defined_names.add(dn)
        result = dispatch_query("list names", model, adapter.index)
        assert "MyRange" in result
        assert "workbook" in result


class TestListTables:
    def test_list_tables_none(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list tables", model, adapter.index)
        assert "Tables: (none)" in result

    def test_list_tables_with_table(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        _setup_simple_data(adapter, model)
        adapter.dispatch_op(parse_op("table add Scores range:A1:C3"), model, log)
        result = dispatch_query("list tables", model, adapter.index)
        assert "Tables" in result
        assert "Scores" in result
        assert "A1:C3" in result


class TestListInvalid:
    def test_list_no_args(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list", model, adapter.index)
        assert "Usage" in result

    def test_list_unknown_subcmd(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("list foobar", model, adapter.index)
        assert "Usage" in result


# ===========================================================================
# Wave 3: find
# ===========================================================================

class TestFindText:
    def test_find_no_args(self, adapter: SheetsAdapter, model: SheetsModel):
        result = dispatch_query("find", model, adapter.index)
        assert "Usage" in result

    def test_find_text_match(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("find Alice", model, adapter.index)
        assert "Alice" in result
        assert "match" in result
        assert "A2" in result

    def test_find_text_case_insensitive(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("find alice", model, adapter.index)
        assert "Alice" in result

    def test_find_text_no_match(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("find Zephyr", model, adapter.index)
        assert "no matches" in result

    def test_find_text_multiple_matches(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 apple"), model, log)
        adapter.dispatch_op(parse_op("set A2 apple_pie"), model, log)
        adapter.dispatch_op(parse_op("set A3 pineapple"), model, log)
        result = dispatch_query("find apple", model, adapter.index)
        assert "3 match" in result

    def test_find_number(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("find 95", model, adapter.index)
        assert "95" in result
        assert "B2" in result

    def test_find_across_sheets(self, adapter: SheetsAdapter):
        model = adapter.create_empty("Multi", {"sheets": "2"})
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 hello"), model, log)
        adapter.dispatch_op(parse_op("sheet activate Sheet2"), model, log)
        adapter.dispatch_op(parse_op("set A1 hello_world"), model, log)
        result = dispatch_query("find hello", model, adapter.index)
        # Should find matches across both sheets
        assert "Sheet1" in result
        assert "Sheet2" in result


class TestFindFormula:
    def test_find_formula_match(self, adapter: SheetsAdapter, model: SheetsModel):
        log = EventLog()
        adapter.dispatch_op(parse_op("set A1 10"), model, log)
        adapter.dispatch_op(parse_op("set A2 =A1*2"), model, log)
        result = dispatch_query("find formula:A1", model, adapter.index)
        assert "=A1*2" in result
        assert "match" in result

    def test_find_formula_no_match(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_simple_data(adapter, model)
        result = dispatch_query("find formula:VLOOKUP", model, adapter.index)
        assert "no matches" in result

    def test_find_formula_case_insensitive(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        # Formulas use =B2+C2+D2, search for lowercase "b"
        result = dispatch_query("find formula:b", model, adapter.index)
        assert "=B" in result

    def test_find_formula_multiple(self, adapter: SheetsAdapter, model: SheetsModel):
        _setup_revenue_sheet(adapter, model)
        # All formulas contain "+C"
        result = dispatch_query("find formula:+C", model, adapter.index)
        assert "3 match" in result


# ===========================================================================
# Wave 3: formatter helpers
# ===========================================================================

class TestFormatterHelpers:
    def test_format_cell_addr(self):
        from fcp_sheets.server.formatter import format_cell_addr
        assert format_cell_addr(1, 1) == "A1"
        assert format_cell_addr(26, 100) == "Z100"
        assert format_cell_addr(27, 1) == "AA1"

    def test_format_range(self):
        from fcp_sheets.server.formatter import format_range
        assert format_range(1, 1, 10, 5) == "A1:E10"

    def test_truncate_list_short(self):
        from fcp_sheets.server.formatter import truncate_list
        assert truncate_list(["a", "b", "c"]) == "a, b, c"

    def test_truncate_list_long(self):
        from fcp_sheets.server.formatter import truncate_list
        items = [str(i) for i in range(20)]
        result = truncate_list(items, max_items=3)
        assert "... +17 more" in result

    def test_format_cell_value_text(self):
        from fcp_sheets.server.formatter import format_cell_value
        assert '"Hello"' in format_cell_value("Hello")

    def test_format_cell_value_formula(self):
        from fcp_sheets.server.formatter import format_cell_value
        assert format_cell_value("=SUM(A1:A10)") == "=SUM(A1:A10)"

    def test_format_cell_value_number(self):
        from fcp_sheets.server.formatter import format_cell_value
        assert format_cell_value(42) == "42"
        assert format_cell_value(3.14) == "3.14"

    def test_format_cell_value_none(self):
        from fcp_sheets.server.formatter import format_cell_value
        assert format_cell_value(None) == "(empty)"

    def test_format_value_type(self):
        from fcp_sheets.server.formatter import format_value_type
        assert format_value_type(None) == "empty"
        assert format_value_type("hello") == "text"
        assert format_value_type("=SUM(A1)") == "formula"
        assert format_value_type(42) == "number"
        assert format_value_type(3.14) == "number"

    def test_format_font(self):
        from openpyxl.styles import Font
        from fcp_sheets.server.formatter import format_font
        font = Font(name="Arial", size=12, bold=True)
        result = format_font(font)
        assert "Arial" in result
        assert "12pt" in result
        assert "bold" in result

    def test_format_fill_none(self):
        from openpyxl.styles import PatternFill
        from fcp_sheets.server.formatter import format_fill
        result = format_fill(PatternFill())
        assert "(none)" in result

    def test_format_alignment_default(self):
        from openpyxl.styles import Alignment
        from fcp_sheets.server.formatter import format_alignment
        result = format_alignment(Alignment())
        assert "(default)" in result

    def test_format_border_none(self):
        from openpyxl.styles import Border
        from fcp_sheets.server.formatter import format_border
        result = format_border(Border())
        assert "(none)" in result
