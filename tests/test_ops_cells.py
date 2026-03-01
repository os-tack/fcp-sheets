"""Tests for cell operations — set, clear verbs."""

from __future__ import annotations

import pytest
from fcp_core import EventLog, ParsedOp

from fcp_sheets.adapter import SheetsAdapter
from fcp_sheets.model.snapshot import SheetsModel
from fcp_sheets.server.ops_cells import op_set, op_clear, _parse_cell_value
from fcp_sheets.server.resolvers import SheetsOpContext


class TestParseValue:
    def test_integer(self):
        assert _parse_cell_value("42") == 42

    def test_float(self):
        assert _parse_cell_value("3.14") == 3.14

    def test_negative_int(self):
        assert _parse_cell_value("-5") == -5

    def test_negative_float(self):
        assert _parse_cell_value("-2.5") == -2.5

    def test_formula(self):
        assert _parse_cell_value("=SUM(A1:A10)") == "=SUM(A1:A10)"

    def test_quoted_string(self):
        assert _parse_cell_value('"Hello"') == "Hello"

    def test_single_quoted(self):
        assert _parse_cell_value("'World'") == "World"

    def test_plain_text(self):
        assert _parse_cell_value("Hello") == "Hello"

    def test_leading_zero_preserved(self):
        """C1: Leading zeros preserved as text."""
        assert _parse_cell_value("01234") == "01234"
        assert _parse_cell_value("007") == "007"

    def test_single_zero(self):
        """Single zero is a number, not leading-zero text."""
        assert _parse_cell_value("0") == 0

    def test_zero_point(self):
        assert _parse_cell_value("0.5") == 0.5


class TestSetVerb:
    def test_set_number(self, ctx: SheetsOpContext, adapter: SheetsAdapter):
        op = ParsedOp(verb="set", positionals=["A1", "42"], raw="set A1 42")
        result = op_set(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).value == 42

    def test_set_text(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="set", positionals=["B2", "Hello"], raw="set B2 Hello")
        result = op_set(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=2, column=2).value == "Hello"

    def test_set_formula(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="set", positionals=["C3", "=SUM(A1:B2)"], raw="set C3 =SUM(A1:B2)")
        result = op_set(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=3, column=3).value == "=SUM(A1:B2)"

    def test_set_with_format(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="set", positionals=["D4", "0.156"], params={"fmt": "0.00%"}, raw="set D4 0.156 fmt:0.00%")
        result = op_set(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=4, column=4)
        assert cell.value == 0.156
        assert cell.number_format == "0.00%"

    def test_set_with_format_alias(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="set", positionals=["A1", "50000"], params={"fmt": "currency"}, raw="set A1 50000 fmt:currency")
        result = op_set(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.value == 50000
        assert cell.number_format == "$#,##0"

    def test_set_missing_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="set", positionals=["A1"], raw="set A1")
        result = op_set(op, ctx)
        assert not result.success

    def test_set_invalid_ref(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="set", positionals=["INVALID", "42"], raw="set INVALID 42")
        result = op_set(op, ctx)
        assert not result.success

    def test_set_updates_index(self, ctx: SheetsOpContext, adapter: SheetsAdapter):
        op = ParsedOp(verb="set", positionals=["E5", "100"], raw="set E5 100")
        op_set(op, ctx)
        bounds = ctx.index.get_bounds("Sheet1")
        assert bounds is not None
        assert bounds[2] >= 5  # max_row includes row 5
        assert bounds[3] >= 5  # max_col includes col 5


class TestSetWithAnchor:
    """C6: Spatial anchor tests for set verb."""

    def test_set_at_bottom_left(self, ctx: SheetsOpContext, adapter: SheetsAdapter):
        # First put some data
        ctx.wb.active.cell(row=1, column=1, value="Header")
        ctx.wb.active.cell(row=2, column=1, value="Data")
        ctx.index.expand_bounds("Sheet1", 1, 1)
        ctx.index.expand_bounds("Sheet1", 2, 1)

        op = ParsedOp(verb="set", positionals=["@bottom_left", "Total"], raw="set @bottom_left Total")
        result = op_set(op, ctx)
        assert result.success
        # @bottom_left should be row 3 (max_row+1), col 1 (min_col)
        assert ctx.wb.active.cell(row=3, column=1).value == "Total"

    def test_set_at_bottom_left_offset(self, ctx: SheetsOpContext, adapter: SheetsAdapter):
        ctx.wb.active.cell(row=1, column=1, value="A")
        ctx.wb.active.cell(row=3, column=2, value="B")
        ctx.index.expand_bounds("Sheet1", 1, 1)
        ctx.index.expand_bounds("Sheet1", 3, 2)

        op = ParsedOp(verb="set", positionals=["@bottom_left+2", "Far"], raw="set @bottom_left+2 Far")
        result = op_set(op, ctx)
        assert result.success
        # @bottom_left+2 = row (3+1+2)=6, col 1
        assert ctx.wb.active.cell(row=6, column=1).value == "Far"


# -- Clear Tests --


class TestClearValues:
    """Clear cell values only."""

    def test_clear_single_cell(self, ctx: SheetsOpContext):
        ctx.wb.active.cell(row=1, column=1, value=42)
        ctx.index.expand_bounds("Sheet1", 1, 1)

        op = ParsedOp(verb="clear", positionals=["A1"], raw="clear A1")
        result = op_clear(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).value is None

    def test_clear_range(self, ctx: SheetsOpContext):
        for r in range(1, 4):
            for c in range(1, 4):
                ctx.wb.active.cell(row=r, column=c, value=r * c)
                ctx.index.expand_bounds("Sheet1", r, c)

        op = ParsedOp(verb="clear", positionals=["A1:C3"], raw="clear A1:C3")
        result = op_clear(op, ctx)
        assert result.success
        assert "9 cells" in result.message
        for r in range(1, 4):
            for c in range(1, 4):
                assert ctx.wb.active.cell(row=r, column=c).value is None

    def test_clear_preserves_formatting(self, ctx: SheetsOpContext):
        """Clear values only should NOT reset formatting."""
        from openpyxl.styles import Font
        cell = ctx.wb.active.cell(row=1, column=1, value="Bold")
        cell.font = Font(bold=True)
        ctx.index.expand_bounds("Sheet1", 1, 1)

        op = ParsedOp(verb="clear", positionals=["A1"], raw="clear A1")
        result = op_clear(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).value is None
        assert ctx.wb.active.cell(row=1, column=1).font.bold is True

    def test_clear_empty_cell(self, ctx: SheetsOpContext):
        """Clearing an already empty cell should succeed."""
        op = ParsedOp(verb="clear", positionals=["A1"], raw="clear A1")
        result = op_clear(op, ctx)
        assert result.success


class TestClearAll:
    """Clear values AND formatting."""

    def test_clear_all_resets_formatting(self, ctx: SheetsOpContext):
        from openpyxl.styles import Font, PatternFill
        cell = ctx.wb.active.cell(row=1, column=1, value="Styled")
        cell.font = Font(bold=True, size=16)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.number_format = "0.00%"
        ctx.index.expand_bounds("Sheet1", 1, 1)

        op = ParsedOp(verb="clear", positionals=["A1", "all"], raw="clear A1 all")
        result = op_clear(op, ctx)
        assert result.success
        assert "values+formatting" in result.message

        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.value is None
        assert cell.font.bold is not True
        assert cell.number_format == "General"

    def test_clear_all_range(self, ctx: SheetsOpContext):
        from openpyxl.styles import Font
        for r in range(1, 3):
            for c in range(1, 3):
                cell = ctx.wb.active.cell(row=r, column=c, value="X")
                cell.font = Font(italic=True)
                ctx.index.expand_bounds("Sheet1", r, c)

        op = ParsedOp(verb="clear", positionals=["A1:B2", "all"], raw="clear A1:B2 all")
        result = op_clear(op, ctx)
        assert result.success
        assert "4 cells" in result.message
        for r in range(1, 3):
            for c in range(1, 3):
                cell = ctx.wb.active.cell(row=r, column=c)
                assert cell.value is None
                assert cell.font.italic is not True


class TestClearErrors:
    """Clear error cases."""

    def test_clear_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="clear", positionals=[], raw="clear")
        result = op_clear(op, ctx)
        assert not result.success
        assert "Usage" in result.message

    def test_clear_invalid_range(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="clear", positionals=["INVALID"], raw="clear INVALID")
        result = op_clear(op, ctx)
        assert not result.success
        assert "Invalid" in result.message or "empty" in result.message


class TestMergedCellProtection:
    """Writing to merged cells should fail gracefully, not crash."""

    def test_set_merged_cell_returns_error(self, ctx: SheetsOpContext):
        """Writing to a non-top-left cell in a merged range should return an error."""
        ws = ctx.active_sheet
        ws.cell(row=1, column=1, value="Title")
        ws.merge_cells("A1:D1")

        # A1 (top-left) should still work
        op_a1 = ParsedOp(verb="set", positionals=["A1", "New Title"], raw='set A1 "New Title"')
        result = op_set(op_a1, ctx)
        assert result.success

        # B1 (merged cell) should return error, not crash
        op_b1 = ParsedOp(verb="set", positionals=["B1", "Bad"], raw='set B1 "Bad"')
        result = op_set(op_b1, ctx)
        assert not result.success
        assert "merged range" in result.message

    def test_set_merged_cell_c1_returns_error(self, ctx: SheetsOpContext):
        """C1 and D1 within A1:D1 merge should also return errors."""
        ws = ctx.active_sheet
        ws.merge_cells("A1:D1")

        op = ParsedOp(verb="set", positionals=["C1", "Bad"], raw='set C1 "Bad"')
        result = op_set(op, ctx)
        assert not result.success
        assert "merged range" in result.message

    def test_data_block_skips_merged_cells(
        self, adapter: SheetsAdapter, model: SheetsModel, log: EventLog,
    ):
        """Data block should skip merged cells with a warning."""
        ws = model.wb.active
        ws.cell(row=1, column=1, value="Title")
        ws.merge_cells("A1:C1")
        adapter.index.active_sheet = ws.title

        # Write data block starting at A1 — row 1 has merged cells
        start = ParsedOp(verb="data", positionals=["A1"], raw="data A1")
        adapter.dispatch_op(start, model, log)

        line1 = ParsedOp(verb="X", positionals=[], raw="H1,H2,H3")
        adapter.dispatch_op(line1, model, log)

        line2 = ParsedOp(verb="a", positionals=[], raw="a,b,c")
        adapter.dispatch_op(line2, model, log)

        end = ParsedOp(verb="data", positionals=["end"], raw="data end")
        result = adapter.dispatch_op(end, model, log)

        assert result.success
        assert "merged cell" in result.message.lower()
        # Row 2 should have data written successfully
        assert ws.cell(row=2, column=1).value == "a"
        assert ws.cell(row=2, column=2).value == "b"


class TestClearViaAdapter:
    """Clear through the adapter dispatch (includes snapshot/undo)."""

    def test_clear_via_dispatch(self, adapter: SheetsAdapter, model: SheetsModel, log: EventLog):
        model.wb.active.cell(row=1, column=1, value=42)
        adapter.index.expand_bounds("Sheet1", 1, 1)

        op = ParsedOp(verb="clear", positionals=["A1"], raw="clear A1")
        result = adapter.dispatch_op(op, model, log)
        assert result.success
        assert model.wb.active.cell(row=1, column=1).value is None
        assert len(log) == 1
