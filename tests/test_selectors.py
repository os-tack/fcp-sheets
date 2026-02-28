"""Tests for selector resolution — @range, @row, @col, @type, @all, @recent, @not, etc."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from fcp_sheets.model.index import SheetIndex
from fcp_sheets.server.resolvers import (
    SheetsOpContext,
    resolve_selectors,
    resolve_target_cells,
)


@pytest.fixture
def populated_ctx() -> SheetsOpContext:
    """Context with data for selector tests.

    Sheet1 has data in A1:D5:
      A1="Name"  B1="Score"  C1="Grade"  D1=100
      A2="Alice" B2=90       C2="A"      D2=95.5
      A3="Bob"   B3=80       C3="B"      D3=85.0
      A4="Carol" B4=70       C4="=B4*1.1" D4=None
      A5=None    B5=None     C5=None     D5=None
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=2, value="Score")
    ws.cell(row=1, column=3, value="Grade")
    ws.cell(row=1, column=4, value=100)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=2, value=90)
    ws.cell(row=2, column=3, value="A")
    ws.cell(row=2, column=4, value=95.5)
    ws.cell(row=3, column=1, value="Bob")
    ws.cell(row=3, column=2, value=80)
    ws.cell(row=3, column=3, value="B")
    ws.cell(row=3, column=4, value=85.0)
    ws.cell(row=4, column=1, value="Carol")
    ws.cell(row=4, column=2, value=70)
    ws.cell(row=4, column=3, value="=B4*1.1")
    # D4 intentionally None
    # Row 5 intentionally all None

    index = SheetIndex()
    # Set bounds to cover A1:D5
    for r in range(1, 6):
        for c in range(1, 5):
            index.expand_bounds("Sheet1", r, c)

    return SheetsOpContext(wb=wb, index=index, named_styles={})


# ── @range ───────────────────────────────────────────────────────────────────


class TestRangeSelector:
    def test_range_a1_d1(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@range:A1:D1"], populated_ctx)
        assert len(cells) == 4
        # All should be row 1
        assert all(row == 1 for _, row, _ in cells)

    def test_range_single_cell(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@range:B2:B2"], populated_ctx)
        assert len(cells) == 1
        ws, row, col = cells[0]
        assert row == 2 and col == 2

    def test_range_a1_d4(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@range:A1:D4"], populated_ctx)
        assert len(cells) == 16  # 4 rows x 4 cols

    def test_range_full_data(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@range:A1:D5"], populated_ctx)
        assert len(cells) == 20  # 5 rows x 4 cols


# ── @row ─────────────────────────────────────────────────────────────────────


class TestRowSelector:
    def test_single_row(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@row:1"], populated_ctx)
        # Should get 4 cells (columns A-D)
        assert len(cells) == 4
        assert all(row == 1 for _, row, _ in cells)

    def test_row_range(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@row:1-3"], populated_ctx)
        # 3 rows x 4 columns = 12 cells
        assert len(cells) == 12
        rows = {row for _, row, _ in cells}
        assert rows == {1, 2, 3}

    def test_row_5_all_empty(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@row:5"], populated_ctx)
        # Row 5 is within bounds (4 cols)
        assert len(cells) == 4


# ── @col ─────────────────────────────────────────────────────────────────────


class TestColSelector:
    def test_single_col(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@col:A"], populated_ctx)
        # Column A with rows 1-5 = 5 cells
        assert len(cells) == 5
        assert all(col == 1 for _, _, col in cells)

    def test_col_range(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@col:A-B"], populated_ctx)
        # 2 columns x 5 rows = 10 cells
        assert len(cells) == 10
        cols = {col for _, _, col in cells}
        assert cols == {1, 2}

    def test_col_d(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@col:D"], populated_ctx)
        assert len(cells) == 5


# ── @type ────────────────────────────────────────────────────────────────────


class TestTypeSelector:
    def test_type_formula(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@type:formula"], populated_ctx)
        # Only C4 has a formula
        assert len(cells) == 1
        ws, row, col = cells[0]
        assert row == 4 and col == 3

    def test_type_number(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@type:number"], populated_ctx)
        # D1=100, B2=90, D2=95.5, B3=80, D3=85.0, B4=70
        assert len(cells) == 6

    def test_type_text(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@type:text"], populated_ctx)
        # A1="Name", B1="Score", C1="Grade", A2="Alice", C2="A",
        # A3="Bob", C3="B", A4="Carol"
        assert len(cells) == 8

    def test_type_empty(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@type:empty"], populated_ctx)
        # D4=None, A5=None, B5=None, C5=None, D5=None = 5 empty cells
        assert len(cells) == 5


# ── @all ─────────────────────────────────────────────────────────────────────


class TestAllSelector:
    def test_all_returns_data_cells(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@all"], populated_ctx)
        # 15 non-None cells (4+4+4+3 in rows 1-4, row 5 all None)
        assert len(cells) == 15

    def test_all_empty_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        index = SheetIndex()
        ctx = SheetsOpContext(wb=wb, index=index, named_styles={})

        cells = resolve_selectors(["@all"], ctx)
        assert len(cells) == 0


# ── @recent ──────────────────────────────────────────────────────────────────


class TestRecentSelector:
    def test_recent_default(self, populated_ctx: SheetsOpContext):
        """@recent returns the last modified range."""
        populated_ctx.index.record_modified("Sheet1", "B2")
        cells = resolve_selectors(["@recent"], populated_ctx)
        assert len(cells) == 1
        _, row, col = cells[0]
        assert row == 2 and col == 2

    def test_recent_n(self, populated_ctx: SheetsOpContext):
        """@recent:3 returns the last 3 modified ranges."""
        populated_ctx.index.record_modified("Sheet1", "A1")
        populated_ctx.index.record_modified("Sheet1", "B2")
        populated_ctx.index.record_modified("Sheet1", "C3")
        cells = resolve_selectors(["@recent:3"], populated_ctx)
        assert len(cells) == 3

    def test_recent_range_format(self, populated_ctx: SheetsOpContext):
        """@recent with range format A1..D4."""
        populated_ctx.index.record_modified("Sheet1", "A1..B2")
        cells = resolve_selectors(["@recent"], populated_ctx)
        # A1:B2 = 4 cells
        assert len(cells) == 4

    def test_recent_no_modifications(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors(["@recent"], populated_ctx)
        assert len(cells) == 0


# ── @not ─────────────────────────────────────────────────────────────────────


class TestNotSelector:
    def test_not_type_formula(self, populated_ctx: SheetsOpContext):
        """@not:type:formula should exclude formula cells from @all."""
        all_cells = resolve_selectors(["@all"], populated_ctx)
        non_formula = resolve_selectors(["@not:type:formula"], populated_ctx)
        formula = resolve_selectors(["@type:formula"], populated_ctx)
        assert len(non_formula) == len(all_cells) - len(formula)

    def test_not_type_empty(self, populated_ctx: SheetsOpContext):
        """@not:type:empty gives cells that have values."""
        non_empty = resolve_selectors(["@not:type:empty"], populated_ctx)
        # All non-None cells
        assert len(non_empty) == 15  # Same as @all since @all filters None

    def test_not_type_number(self, populated_ctx: SheetsOpContext):
        not_number = resolve_selectors(["@not:type:number"], populated_ctx)
        numbers = resolve_selectors(["@type:number"], populated_ctx)
        all_cells = resolve_selectors(["@all"], populated_ctx)
        # not_number = all - number (note: @all only includes non-None)
        # but @not works from universe = @all
        assert len(not_number) == len(all_cells) - len(numbers)


# ── Combined selectors (intersection) ───────────────────────────────────────


class TestCombinedSelectors:
    def test_row_and_col_intersection(self, populated_ctx: SheetsOpContext):
        """@row:1 AND @col:A should give just A1."""
        cells = resolve_selectors(["@row:1", "@col:A"], populated_ctx)
        assert len(cells) == 1
        _, row, col = cells[0]
        assert row == 1 and col == 1

    def test_range_and_type(self, populated_ctx: SheetsOpContext):
        """@range:A1:D2 AND @type:number."""
        cells = resolve_selectors(["@range:A1:D2", "@type:number"], populated_ctx)
        # Row 1-2, number cells: D1=100, B2=90, D2=95.5
        assert len(cells) == 3

    def test_row_range_and_type_text(self, populated_ctx: SheetsOpContext):
        """@row:1-2 AND @type:text."""
        cells = resolve_selectors(["@row:1-2", "@type:text"], populated_ctx)
        # Row 1: Name, Score, Grade (3 text); Row 2: Alice, A (2 text) = 5
        assert len(cells) == 5

    def test_three_way_intersection(self, populated_ctx: SheetsOpContext):
        """@row:2 AND @col:A-B AND @type:text."""
        cells = resolve_selectors(["@row:2", "@col:A-B", "@type:text"], populated_ctx)
        # Row 2, Cols A-B: A2="Alice" (text), B2=90 (number)
        assert len(cells) == 1
        _, row, col = cells[0]
        assert row == 2 and col == 1  # A2="Alice"


# ── resolve_target_cells ─────────────────────────────────────────────────────


class TestResolveTargetCells:
    def test_with_range_positional(self, populated_ctx: SheetsOpContext):
        cells = resolve_target_cells(["A1:B2"], [], populated_ctx)
        assert len(cells) == 4

    def test_with_selector(self, populated_ctx: SheetsOpContext):
        cells = resolve_target_cells([], ["@row:1"], populated_ctx)
        assert len(cells) == 4

    def test_selector_as_positional(self, populated_ctx: SheetsOpContext):
        """If first positional starts with @, treat as selector."""
        cells = resolve_target_cells(["@range:A1:B2"], [], populated_ctx)
        assert len(cells) == 4

    def test_no_args(self, populated_ctx: SheetsOpContext):
        cells = resolve_target_cells([], [], populated_ctx)
        assert len(cells) == 0

    def test_empty_selectors_list(self, populated_ctx: SheetsOpContext):
        cells = resolve_selectors([], populated_ctx)
        assert len(cells) == 0
