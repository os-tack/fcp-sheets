"""Tests for style operation handlers — style, border, define-style, apply-style."""

from __future__ import annotations

import pytest
from copy import copy
from fcp_core import ParsedOp

from fcp_sheets.adapter import SheetsAdapter
from fcp_sheets.model.snapshot import SheetsModel
from fcp_sheets.server.ops_style import op_style, op_border, op_define_style, op_apply_style
from fcp_sheets.server.resolvers import SheetsOpContext


# ── Style verb ───────────────────────────────────────────────────────────────


class TestStyleBold:
    def test_bold_single_cell(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="style", positionals=["A1", "bold"], raw="style A1 bold")
        result = op_style(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.font.bold is True

    def test_bold_range(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="style", positionals=["A1:B2", "bold"], raw="style A1:B2 bold")
        result = op_style(op, ctx)
        assert result.success
        for r in (1, 2):
            for c in (1, 2):
                assert ctx.wb.active.cell(row=r, column=c).font.bold is True


class TestStyleItalic:
    def test_italic(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="style", positionals=["C3", "italic"], raw="style C3 italic")
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=3, column=3).font.italic is True


class TestStyleUnderline:
    def test_underline(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="style", positionals=["A1", "underline"], raw="style A1 underline")
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).font.underline == "single"


class TestStyleStrike:
    def test_strikethrough(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="style", positionals=["A1", "strike"], raw="style A1 strike")
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).font.strike is True


class TestStyleFont:
    def test_font_name(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"font": "Helvetica"}, raw="style A1 font:Helvetica",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).font.name == "Helvetica"

    def test_font_size(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"size": "14"}, raw="style A1 size:14",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).font.size == 14.0


class TestStyleColor:
    def test_hex_color(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"color": "#FF0000"}, raw="style A1 color:#FF0000",
        )
        result = op_style(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        # openpyxl stores color as a Color object; font.color can be checked
        assert cell.font.color is not None

    def test_named_color(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"color": "blue"}, raw="style A1 color:blue",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).font.color is not None


class TestStyleFill:
    def test_fill_color(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"fill": "#FFFF00"}, raw="style A1 fill:#FFFF00",
        )
        result = op_style(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.fill.start_color.rgb is not None
        assert cell.fill.fill_type == "solid"

    def test_fill_named_color(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"fill": "green"}, raw="style A1 fill:green",
        )
        result = op_style(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.fill.fill_type == "solid"


class TestStyleAlignment:
    def test_align_center(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"align": "center"}, raw="style A1 align:center",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).alignment.horizontal == "center"

    def test_align_right(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"align": "right"}, raw="style A1 align:right",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).alignment.horizontal == "right"

    def test_align_left(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"align": "left"}, raw="style A1 align:left",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).alignment.horizontal == "left"

    def test_valign_top(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"valign": "top"}, raw="style A1 valign:top",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).alignment.vertical == "top"

    def test_valign_middle(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"valign": "middle"}, raw="style A1 valign:middle",
        )
        result = op_style(op, ctx)
        assert result.success
        # "middle" is mapped to "center" for openpyxl compatibility
        assert ctx.wb.active.cell(row=1, column=1).alignment.vertical == "center"


class TestStyleWrap:
    def test_wrap_text(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="style", positionals=["A1", "wrap"], raw="style A1 wrap")
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).alignment.wrap_text is True


class TestStyleIndent:
    def test_indent(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"indent": "2"}, raw="style A1 indent:2",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).alignment.indent == 2


class TestStyleRotate:
    def test_rotate(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"rotate": "45"}, raw="style A1 rotate:45",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).alignment.text_rotation == 45


class TestStyleFormat:
    def test_number_format_raw(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"fmt": "0.00%"}, raw="style A1 fmt:0.00%",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).number_format == "0.00%"

    def test_number_format_alias(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1"],
            params={"fmt": "currency"}, raw="style A1 fmt:currency",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).number_format == "$#,##0"


class TestStyleCombined:
    def test_bold_italic_with_font(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1", "bold", "italic"],
            params={"font": "Arial", "size": "12"},
            raw="style A1 bold italic font:Arial size:12",
        )
        result = op_style(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.font.name == "Arial"
        assert cell.font.size == 12.0

    def test_full_styling(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="style", positionals=["A1:B2", "bold", "wrap"],
            params={
                "font": "Calibri", "size": "11", "color": "#333333",
                "fill": "#EEEEEE", "align": "center", "valign": "middle",
                "fmt": "number2",
            },
            raw="style A1:B2 bold wrap font:Calibri size:11 color:#333333 fill:#EEEEEE align:center valign:middle fmt:number2",
        )
        result = op_style(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.name == "Calibri"
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.wrap_text is True
        assert cell.number_format == "0.00"

    def test_preserves_existing_font_attrs(self, ctx: SheetsOpContext):
        """When applying bold, existing font name should be preserved."""
        from openpyxl.styles import Font
        cell = ctx.wb.active.cell(row=1, column=1)
        cell.font = Font(name="Courier New", size=16)

        op = ParsedOp(verb="style", positionals=["A1", "bold"], raw="style A1 bold")
        op_style(op, ctx)

        assert cell.font.bold is True
        assert cell.font.name == "Courier New"
        assert cell.font.size == 16.0


class TestStyleErrors:
    def test_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="style", positionals=[], raw="style")
        result = op_style(op, ctx)
        assert not result.success

    def test_no_cells_resolved(self, ctx: SheetsOpContext):
        """Style with selector that matches no cells."""
        op = ParsedOp(
            verb="style", positionals=["bold"],
            selectors=["@type:formula"],
            raw="style bold @type:formula",
        )
        result = op_style(op, ctx)
        assert not result.success


class TestStyleWithSelectors:
    def test_style_with_range_selector(self, ctx: SheetsOpContext):
        """Style cells using @range selector."""
        ws = ctx.wb.active
        ws.cell(row=1, column=1, value="A")
        ws.cell(row=1, column=2, value="B")
        ctx.index.expand_bounds("Sheet1", 1, 1)
        ctx.index.expand_bounds("Sheet1", 1, 2)

        op = ParsedOp(
            verb="style", positionals=["bold"],
            selectors=["@range:A1:B1"],
            raw="style bold @range:A1:B1",
        )
        result = op_style(op, ctx)
        assert result.success
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=1, column=2).font.bold is True


# ── Border verb ──────────────────────────────────────────────────────────────


class TestBorderAll:
    def test_border_all(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1:B2", "all"],
            raw="border A1:B2 all",
        )
        result = op_border(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"

    def test_border_all_cells_in_range(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1:C3", "all"],
            raw="border A1:C3 all",
        )
        result = op_border(op, ctx)
        assert result.success
        for r in range(1, 4):
            for c in range(1, 4):
                cell = ctx.wb.active.cell(row=r, column=c)
                assert cell.border.top.style is not None


class TestBorderOutline:
    def test_border_outline(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1:C3", "outline"],
            raw="border A1:C3 outline",
        )
        result = op_border(op, ctx)
        assert result.success
        ws = ctx.wb.active
        # Top-left corner: top and left borders
        assert ws.cell(row=1, column=1).border.top.style == "thin"
        assert ws.cell(row=1, column=1).border.left.style == "thin"
        # Bottom-right corner: bottom and right borders
        assert ws.cell(row=3, column=3).border.bottom.style == "thin"
        assert ws.cell(row=3, column=3).border.right.style == "thin"
        # Interior cell: no outline borders
        assert ws.cell(row=2, column=2).border.top.style is None
        assert ws.cell(row=2, column=2).border.left.style is None


class TestBorderSides:
    def test_border_top(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "top"],
            raw="border A1 top",
        )
        result = op_border(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).border.top.style == "thin"

    def test_border_bottom(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "bottom"],
            raw="border A1 bottom",
        )
        result = op_border(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).border.bottom.style == "thin"

    def test_border_left(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "left"],
            raw="border A1 left",
        )
        result = op_border(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).border.left.style == "thin"

    def test_border_right(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "right"],
            raw="border A1 right",
        )
        result = op_border(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).border.right.style == "thin"


class TestBorderLineStyle:
    def test_medium_border(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "all"],
            params={"line": "medium"}, raw="border A1 all line:medium",
        )
        result = op_border(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).border.top.style == "medium"

    def test_thick_border(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "all"],
            params={"line": "thick"}, raw="border A1 all line:thick",
        )
        result = op_border(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).border.top.style == "thick"

    def test_dashed_border(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "all"],
            params={"line": "dashed"}, raw="border A1 all line:dashed",
        )
        result = op_border(op, ctx)
        assert result.success
        assert ctx.wb.active.cell(row=1, column=1).border.top.style == "dashed"

    def test_invalid_line_style(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "all"],
            params={"line": "zigzag"}, raw="border A1 all line:zigzag",
        )
        result = op_border(op, ctx)
        assert not result.success


class TestBorderColor:
    def test_border_with_color(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="border", positionals=["A1", "all"],
            params={"color": "#FF0000"}, raw="border A1 all color:#FF0000",
        )
        result = op_border(op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.border.top.color is not None


class TestBorderErrors:
    def test_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="border", positionals=[], raw="border")
        result = op_border(op, ctx)
        assert not result.success

    def test_missing_sides(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="border", positionals=["A1"], raw="border A1")
        result = op_border(op, ctx)
        assert not result.success


# ── Define-style / apply-style ───────────────────────────────────────────────


class TestDefineStyle:
    def test_define_basic_style(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="define-style", positionals=["header", "bold"],
            params={"font": "Arial", "size": "14", "fill": "#4472C4", "color": "#FFFFFF"},
            raw="define-style header bold font:Arial size:14 fill:#4472C4 color:#FFFFFF",
        )
        result = op_define_style(op, ctx)
        assert result.success
        assert "header" in ctx.named_styles
        assert ctx.named_styles["header"]["bold"] is True
        assert ctx.named_styles["header"]["font"] == "Arial"

    def test_define_style_no_name(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="define-style", positionals=[], raw="define-style")
        result = op_define_style(op, ctx)
        assert not result.success

    def test_define_style_with_flags(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="define-style", positionals=["emphasis", "bold", "italic"],
            raw="define-style emphasis bold italic",
        )
        result = op_define_style(op, ctx)
        assert result.success
        assert ctx.named_styles["emphasis"]["bold"] is True
        assert ctx.named_styles["emphasis"]["italic"] is True


class TestApplyStyle:
    def test_apply_defined_style(self, ctx: SheetsOpContext):
        # First define a style
        define_op = ParsedOp(
            verb="define-style", positionals=["header", "bold"],
            params={"fill": "#4472C4"},
            raw="define-style header bold fill:#4472C4",
        )
        op_define_style(define_op, ctx)

        # Apply it
        apply_op = ParsedOp(
            verb="apply-style", positionals=["header", "A1:C1"],
            raw="apply-style header A1:C1",
        )
        result = op_apply_style(apply_op, ctx)
        assert result.success
        cell = ctx.wb.active.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.fill.fill_type == "solid"

    def test_apply_unknown_style(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="apply-style", positionals=["nonexistent", "A1"],
            raw="apply-style nonexistent A1",
        )
        result = op_apply_style(op, ctx)
        assert not result.success

    def test_apply_style_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="apply-style", positionals=[], raw="apply-style")
        result = op_apply_style(op, ctx)
        assert not result.success

    def test_apply_style_with_fmt(self, ctx: SheetsOpContext):
        """Define and apply a style with number format."""
        define_op = ParsedOp(
            verb="define-style", positionals=["money"],
            params={"fmt": "currency2", "bold": "true"},
            raw="define-style money fmt:currency2 bold:true",
        )
        op_define_style(define_op, ctx)

        apply_op = ParsedOp(
            verb="apply-style", positionals=["money", "B2"],
            raw="apply-style money B2",
        )
        result = op_apply_style(apply_op, ctx)
        assert result.success
